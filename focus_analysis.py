#!/usr/bin/env python3
"""
重点分析 — 冲三档 + 差2单/20min升档
按公司重点标准输出可跟进名单
"""
import os, re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system('pip3 install openpyxl --break-system-packages >/dev/null 2>&1')
    import openpyxl

LEVELS = [
    {'level': 1, 'orders': 18, 'hours': 6,  'label': '标准1:6h/18单'},
    {'level': 2, 'orders': 28, 'hours': 6,  'label': '标准2:6h/28单'},
    {'level': 3, 'orders': 35, 'hours': 6,  'label': '标准3:6h/35单'},
    {'level': 4, 'orders': 42, 'hours': 7,  'label': '标准4:7h/42单'},
    {'level': 5, 'orders': 53, 'hours': 8,  'label': '标准5:8h/53单'},
]

def read_transfer_data():
    """读取最新运单数据，返回 {骑手名称: 改派单量+无效单量} 字典
    
    无效单定义：运单状态 != '配送成功' 且 != '配送中'
    与监控表的 '运单明细!A:A（无效单量）+ 运单明细!B:B（改派单量）' 保持一致
    """
    files = sorted(Path.home().glob('Downloads/*运单*.xlsx'),
                   key=lambda f: os.path.getmtime(f), reverse=True)
    files = [f for f in files if not f.name.startswith('.') and not f.name.startswith('~') and f.stat().st_size > 10000]
    if not files:
        return {}
    try:
        wb = openpyxl.load_workbook(str(files[0]), data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        transfer_col = None
        status_col = None
        name_col = None
        for i, h in enumerate(headers):
            if h and '是否发生改派' in str(h):
                transfer_col = i + 1
            if h and '运单状态' in str(h):
                status_col = i + 1
            if h and '骑手名称' in str(h):
                name_col = i + 1
        if not name_col:
            wb.close()
            return {}
        result = {}
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            if not name:
                continue
            # 改派扣减
            if transfer_col:
                val = ws.cell(r, transfer_col).value
                if val is not None and str(val).strip() == '是':
                    result[name] = result.get(name, 0) + 1
            # 无效单扣减（运单状态不是配送成功/配送中）
            if status_col:
                s_val = str(ws.cell(r, status_col).value or '')
                if s_val not in ('配送成功', '配送中', ''):
                    result[name] = result.get(name, 0) + 1
        wb.close()
        return result
    except Exception:
        return {}


def fmt_time(h):
    if h is None or h == 0: return '0分钟'
    total_m = round(h * 60)
    hrs, mins = divmod(total_m, 60)
    if hrs == 0: return f'{mins}分钟'
    if mins == 0: return f'{hrs}小时'
    return f'{hrs}小时{mins}分钟'

def parse_time(t):
    if not t or t == '-' or t is None: return 0.0
    # 纯数字→秒（新格式：全天有效在线时长/时段有效在线时长以秒存储）
    if isinstance(t, (int, float)):
        return t / 3600.0
    t = str(t).strip()
    h = 0.0; m = 0.0
    g = re.search(r'(\d+)小时', t)
    if g: h = float(g.group(1))
    g = re.search(r'(\d+)分钟', t)
    if g: m = float(g.group(1))
    if not h and not m:
        g = re.search(r'^(\d+):(\d+)$', t)
        if g: h, m = float(g.group(1)), float(g.group(2))
    return h + m / 60

def calc_level(orders, hours):
    order_lv = 0; hours_lv = 0
    for lv in LEVELS:
        if orders >= lv['orders']: order_lv = lv['level']
        if hours >= lv['hours']:   hours_lv = lv['level']
    return min(order_lv, hours_lv) if order_lv and hours_lv else 0

# 时段时间定义（与 generate_report.py 保持一致）
PERIOD_TIMES = {
    '凌晨1': (0, 2), '凌晨2': (2, 4), '凌晨3': (4, 6),
    '早餐1': (6, 8), '早餐2': (8, 10.5),
    '午高峰': (10.5, 13.5),
    '下午茶1': (13.5, 15.5), '下午茶2': (15.5, 17.5),
    '晚高峰': (17.5, 20),
    '夜宵1': (20, 22), '夜宵2': (22, 24),
}

def _current_hour():
    """获取当前小时数（含分钟）"""
    now = datetime.now()
    return now.hour + now.minute / 60

def get_remaining_periods(schedule_str, current_hour=None):
    """解析排班时段，返回仍未结束的班次列表"""
    if current_hour is None:
        current_hour = _current_hour()
    if not schedule_str:
        return []
    periods = [p.strip() for p in schedule_str.replace('、', ',').replace('，', ',').split(',')]
    remaining = []
    for p in periods:
        p = p.strip()
        if p in PERIOD_TIMES:
            start, end = PERIOD_TIMES[p]
            if end > current_hour:
                remaining.append(p)
    return remaining

def estimate_achieve(current_orders, current_online, target_orders, remaining_periods, delivering=0, current_hour=None, target_hours=None):
    """根据剩余班次预估能否达成目标（含配送中单量）
    target_hours 可选，指定时长目标（如基线6h）
    """
    if current_hour is None:
        current_hour = _current_hour()
    if not remaining_periods:
        return '已无班次'
    effective_orders = current_orders + delivering
    gap = target_orders - effective_orders
    time_gap = (target_hours - current_online) if target_hours else 0
    if gap <= 0 and time_gap <= 0:
        return '已达成'
    # 计算剩余总时长
    remaining_hours = 0
    for p in remaining_periods:
        if p in PERIOD_TIMES:
            start, end = PERIOD_TIMES[p]
            remaining_hours += end - max(current_hour, start)
    # 根据当前效率预估
    avg_per_hour = current_orders / current_online if current_online > 0 else 0
    predicted_extra = remaining_hours * max(avg_per_hour, 0)
    note = f'({fmt_time(remaining_hours)})'
    if delivering > 0:
        note += f'  含配送中{delivering:.0f}单'
    # 同时判断单量和时长
    order_ok = gap <= 0 or predicted_extra >= gap * 0.8
    time_ok = time_gap <= 0 or remaining_hours >= time_gap * 0.8
    if order_ok and time_ok:
        return f'可能{note}'
    else:
        return f'困难{note}'

def analyze(filepath=None, transfer_data=None, transfer_only_data=None):
    if not filepath:
        files = sorted(Path.home().glob('Downloads/*出勤*.xlsx'),
                       key=lambda f: os.path.getmtime(f), reverse=True)
        files = [f for f in files if not f.name.startswith('.') and not f.name.startswith('~') and f.stat().st_size > 10000]
        if not files:
            print("❌ 未找到出勤文件")
            return
        filepath = files[0]

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def find_col(targets, blacklist=None):
        for t in targets:
            for i, h in enumerate(headers):
                if h and t in str(h) and not (blacklist and any(b in str(h) for b in blacklist)):
                    return i + 1
        return None

    col = {
        '站点': find_col(['站点名称']),
        '姓名': find_col(['姓名', '骑手姓名']),
        '时段': find_col(['时段'], blacklist=['排班']),
        '排班': find_col(['排班时段', '排班状态', '是否排班']),
        '排班时段': find_col(['排班时段']),
        '完单': find_col(['全天完单量', '完单量'], blacklist=['目标']),
        '时长': find_col(['全天有效在线时长', '有效在线时长'], blacklist=['目标']),
        '配送中': find_col(['配送中单量']),
        '工作': find_col(['工作状态']),
    }

    if not col['姓名']:
        print(f"❌ 无法识别列: {headers[:20]}")
        return

    riders = {}; seen = set()
    is_period_col = col.get('时段') is not None  # 有时段列=明细格式，无时段列=汇总格式
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col['姓名']).value
        if not name: continue
        if name in seen: continue

        if is_period_col:
            period_v = ws.cell(r, col['时段']).value
            # 时段明细格式：跳过非「全天」的行（各时段明细），只取全天汇总
            if not period_v or str(period_v).strip() != '全天': continue
        # 无时段列：每行都是骑手汇总数据，全部处理
        seen.add(name)

        try: orders = float(ws.cell(r, col['完单']).value or 0)
        except: orders = 0

        # 改派+无效扣减：有效完成单 = 全天完单量 - 改派单量 - 无效单量
        if transfer_data is None:
            transfer_data = read_transfer_data()[0]
        transfer_orders = transfer_data.get(name, 0)
        effective_orders = max(0, orders - transfer_orders)

        online = parse_time(ws.cell(r, col['时长']).value)

        schedule = str(ws.cell(r, col['排班']).value or '') if col['排班'] else '正常'
        if schedule.strip() in ['', '休息', '未排班', '请假']: continue

        # 读取排班时段（后续判断用）
        shift_schedule = str(ws.cell(r, col['排班时段']).value or '') if col.get('排班时段') else ''

        delivering_raw = ws.cell(r, col['配送中']).value if col['配送中'] else 0
        try: delivering = float(delivering_raw or 0)
        except: delivering = 0
        work_status = str(ws.cell(r, col['工作']).value or '') if col['工作'] else ''
        level = calc_level(effective_orders, online)
        riders[name] = {'name': name, 'orders': orders, 'online': round(online, 2), 'level': level, 'delivering': delivering, 'work_status': work_status, 'shift_schedule': shift_schedule, 'effective_orders': effective_orders, 'transfer_orders': transfer_orders, 'transfer_only_orders': (transfer_only_data or {}).get(name, 0)}

    wb.close()

    # ===== 基线过滤：6h/18单 =====
    BASELINE_HOURS = 6
    BASELINE_ORDERS = 18
    baseline_riders = {n: r for n, r in riders.items() if r['orders'] >= BASELINE_ORDERS and r['online'] >= BASELINE_HOURS}
    below_riders = {n: r for n, r in riders.items() if not (r['orders'] >= BASELINE_ORDERS and r['online'] >= BASELINE_HOURS)}

    # ===== 1. 冲三档 =====
    print(f"\n{'='*60}")
    print(f"🎯 重点分析 | 差2单/20min升档 + 冲三档")
    print(f"   文件: {Path(filepath).name}")
    print(f"   基线过滤：≥{BASELINE_HOURS}h & ≥{BASELINE_ORDERS}单 | 基线内{len(baseline_riders)}人 低于基线{len(below_riders)}人")
    print(f"{'='*60}")

    print(f"\n📌 【冲三档】标准2→标准3（6h/35单）差2单或20min内：")
    print(f"{'骑手':<8} {'实完单':>6} {'实在线':>6} {'差单':>4} {'差时':>5} {'升级条件'}")
    print(f"{'─'*45}")
    lv3_candidates = []
    for r in baseline_riders.values():
        if r['level'] == 2:
            eff_o = r.get('effective_orders', r['orders'])
            gap_o = LEVELS[2]['orders'] - eff_o  # 标准3: 35单
            gap_h = round(LEVELS[2]['hours'] - r['online'], 1)
            if (0 < gap_o <= 2) or (0 < gap_h <= 0.33):
                remaining_shifts = get_remaining_periods(r.get('shift_schedule', ''))
                can_achieve = estimate_achieve(eff_o, r['online'], LEVELS[2]['orders'], remaining_shifts, r.get('delivering', 0))
                r['remaining_shifts'] = '、'.join(remaining_shifts) if remaining_shifts else '无'
                r['can_achieve'] = can_achieve
                lv3_candidates.append(r)
                parts = []
                if 0 < gap_o <= 2: parts.append(f"差{gap_o:.0f}单")
                if 0 < gap_h <= 0.33: parts.append(f"差{fmt_time(gap_h)}")
                print(f"{r['name']:<8} {eff_o:>4.0f}单 {fmt_time(r['online'])} {gap_o:>3.0f} {fmt_time(gap_h)} {' '.join(parts)}")
    if not lv3_candidates: print("  （无）")

    # ===== 2. 全部升档机会（基线内） =====
    print(f"\n📌 【全部升档机会】基线内各档位差2单或20min内升档：")
    print(f"{'骑手':<8} {'当前':>6} {'实完单':>6} {'实在线':>6} {'目标':>6} {'差距'}")
    print(f"{'─'*45}")
    all_candidates = []
    for r in baseline_riders.values():
        lv = r['level']
        if lv >= 5: continue
        if lv == 0:
            t_orders, t_hours, t_label = LEVELS[0]['orders'], LEVELS[0]['hours'], '标准1'
        else:
            t_orders, t_hours = LEVELS[lv]['orders'], LEVELS[lv]['hours']
            t_label = f"标准{LEVELS[lv]['level']}"
        eff_o = r.get('effective_orders', r['orders'])
        gap_o = t_orders - eff_o
        gap_h = round(t_hours - r['online'], 1)
        if (0 < gap_o <= 2) or (0 < gap_h <= 0.33):
            remaining_shifts = get_remaining_periods(r.get('shift_schedule', ''))
            can_achieve = estimate_achieve(eff_o, r['online'], t_orders, remaining_shifts, r.get('delivering', 0))
            r['remaining_shifts'] = '、'.join(remaining_shifts) if remaining_shifts else '无'
            r['can_achieve'] = can_achieve
            all_candidates.append({**r, 'target': t_label, 'gap_o': gap_o, 'gap_h': gap_h})
            parts = []
            if 0 < gap_o <= 2: parts.append(f"差{gap_o:.0f}单")
            if 0 < gap_h <= 0.33: parts.append(f"差{fmt_time(gap_h)}")
            print(f"{r['name']:<8} {'未达标' if lv==0 else f'标准{lv}':>6} {eff_o:>4.0f}单 {fmt_time(r['online'])} {t_label:>6} {' '.join(parts)}")

    # ===== 2b. 低于基线的升档机会 =====
    below_candidates = []
    for r in below_riders.values():
        lv = r['level']
        if lv >= 5: continue
        if lv == 0:
            t_orders, t_hours, t_label = LEVELS[0]['orders'], LEVELS[0]['hours'], '标准1'
        else:
            t_orders, t_hours = LEVELS[lv]['orders'], LEVELS[lv]['hours']
            t_label = f"标准{LEVELS[lv]['level']}"
        eff_o = r.get('effective_orders', r['orders'])
        gap_o = t_orders - eff_o
        gap_h = round(t_hours - r['online'], 1)
        if (0 < gap_o <= 2) or (0 < gap_h <= 0.33):
            below_candidates.append({**r, 'target': t_label, 'gap_o': gap_o, 'gap_h': gap_h})
    if below_candidates:
        print(f"\n📌 【低于基线升档机会】{len(below_candidates)}人（数据参考）:")
        for r in sorted(below_candidates, key=lambda x: (-x['effective_orders'] if x.get('effective_orders') else -x['orders'])):
            parts = []
            if 0 < r['gap_o'] <= 2: parts.append(f"差{r['gap_o']:.0f}单")
            if 0 < r['gap_h'] <= 0.33: parts.append(f"差{fmt_time(r['gap_h'])}")
            eff_o = r.get('effective_orders', r['orders'])
            print(f"    {r['name']:<8} {eff_o:>4.0f}单 {fmt_time(r['online'])} → {r['target']} {' '.join(parts)}")

    # ===== 3. 汇总 =====
    lv3_plus = [r for r in baseline_riders.values() if r['level'] >= 3]
    print(f"\n📊 当前标准3以上（基线内）：{len(lv3_plus)}人")
    print(f"📊 冲三档机会（基线内）：{len(lv3_candidates)}人")
    print(f"📊 全部升档机会（基线内）：{len(all_candidates)}人")
    print(f"📊 潜力可达（标准3以上+冲三档）：{len(lv3_plus) + len(lv3_candidates)}人")
    print(f"📊 低于基线骑手：{len(below_riders)}人")

    # 扣减汇总（仅显示改派扣减）
    transfer_summary = {}
    for r in riders.values():
        tf_only = r.get('transfer_only_orders', 0)
        if tf_only > 0:
            transfer_summary[r['name']] = tf_only
    if transfer_summary:
        print(f"📋 改派扣减: {sum(transfer_summary.values())}单")
        for name, cnt in sorted(transfer_summary.items(), key=lambda x: -x[1]):
            print(f"    {name}: {cnt}单")

    # ===== 4. 低于基线（未达成一档）明细 =====
    below_details = {}
    for n, r in below_riders.items():
        eff_o = r.get('effective_orders', r['orders'])
        gap_o = max(0, LEVELS[0]['orders'] - eff_o)
        gap_h = max(0, LEVELS[0]['hours'] - r['online'])
        remaining_shifts = get_remaining_periods(r.get('shift_schedule', ''))
        can_achieve = estimate_achieve(eff_o, r['online'], LEVELS[0]['orders'], remaining_shifts, r.get('delivering', 0), target_hours=LEVELS[0]['hours'])
        below_details[n] = {
            'name': n,
            'orders': r['orders'],
            'online': r['online'],
            'effective_orders': eff_o,
            'transfer_orders': r.get('transfer_only_orders', 0),
            'delivering': r.get('delivering', 0),
            'work_status': r.get('work_status', ''),
            'gap_orders': gap_o,
            'gap_hours': gap_h,
            'remaining_shifts': '、'.join(remaining_shifts) if remaining_shifts else '无',
            'can_achieve': can_achieve,
            'shift_schedule': r.get('shift_schedule', ''),
        }
    if below_details:
        print(f"\n📌 【未达成一档】低于标准1（<6h 或 <18单）: {len(below_details)}人")
        for r in sorted(below_details.values(), key=lambda x: (-x['effective_orders'], -x['online'])):
            print(f"    {r['name']:<8} {r['effective_orders']:>4.0f}单 {fmt_time(r['online'])}  {'扣'+str(r['transfer_orders'])+'单' if r['transfer_orders']>0 else ''}  → {r['can_achieve']}")

    return {'lv3_candidates': lv3_candidates, 'all_candidates': all_candidates, 'lv3_plus': len(lv3_plus), 'below_count': len(below_riders), 'below_details': below_details}

if __name__ == '__main__':
    import sys
    analyze(sys.argv[1] if len(sys.argv) > 1 else None)

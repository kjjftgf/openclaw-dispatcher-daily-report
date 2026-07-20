#!/usr/bin/env python3
"""
# VERSION: v20260618_2348 | 预测列简化版（是/否）
# 版本历史: 每次修改前自动备份到 versions/ 目录
配送日报生成器 — 合并版
WPS 适配配色 + 桌面输出 + 三大板块
"""
import sys, os, re
from datetime import datetime
from pathlib import Path
from importlib import util as import_util

# ── 从合并技能目录加载分析模块 ──
SKILL_DIR = Path.home() / '.openclaw/skills/dispatcher-data-analysis'

def import_mod(name, path):
    spec = import_util.spec_from_file_location(name, path)
    mod = import_util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

att_mod = import_mod('att', str(SKILL_DIR / 'scripts/attendance_analysis.py'))
lv_mod  = import_mod('lv',  str(SKILL_DIR / 'scripts/level_analysis.py'))
fc_mod  = import_mod('fc',  str(SKILL_DIR / 'scripts/focus_analysis.py'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from copy import copy
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════
# 🎨 WPS 适配配色方案
# ═══════════════════════════════════════
C_PRIMARY   = '1A365D'
C_BLUE      = '3182CE'
C_WHITE     = 'FFFFFF'
C_BG        = 'F7FAFC'
C_ALT       = 'EDF2F7'
C_HEADER_BG = '1A365D'
C_BORDER    = 'CBD5E0'
C_TEXT      = '2D3748'
C_MUTED     = '718096'

C_PASS      = '38A169'
C_PASS_BG   = 'F0FFF4'
C_FAIL      = 'E53E3E'
C_FAIL_BG   = 'FFF5F5'
C_WARN      = 'D69E2E'
C_WARN_BG   = 'FFFAF0'

# ── 填充 ──
FILL_HEADER  = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type='solid')
FILL_BG      = PatternFill(start_color=C_BG, end_color=C_BG, fill_type='solid')
FILL_ALT     = PatternFill(start_color=C_ALT, end_color=C_ALT, fill_type='solid')
FILL_PASS    = PatternFill(start_color=C_PASS_BG, end_color=C_PASS_BG, fill_type='solid')
FILL_FAIL    = PatternFill(start_color=C_FAIL_BG, end_color=C_FAIL_BG, fill_type='solid')
FONT_FAIL    = Font(size=10, color=C_FAIL, name='PingFang SC')
FILL_WARN    = PatternFill(start_color=C_WARN_BG, end_color=C_WARN_BG, fill_type='solid')
FILL_SECTION = PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid')
FILL_UNSCHED = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')  # 非排班达成：浅蓝底色

# ── 字体 ──
FONT_TITLE = Font(bold=True, size=14, color=C_WHITE, name='PingFang SC')
FONT_HDR   = Font(bold=True, size=10, color=C_WHITE, name='PingFang SC')
FONT_DATA  = Font(size=10, color=C_TEXT, name='PingFang SC')
FONT_BOLD  = Font(size=10, color=C_TEXT, bold=True, name='PingFang SC')
FONT_MUTED = Font(size=10, color=C_MUTED, name='PingFang SC')
FONT_PASS  = Font(size=10, color=C_PASS, name='PingFang SC')
FONT_FAIL  = Font(size=10, color=C_FAIL, name='PingFang SC')
FONT_WARN  = Font(size=10, color=C_WARN, name='PingFang SC')
FONT_SECTION = Font(size=11, color=C_PRIMARY, bold=True, name='PingFang SC')

# ── 边框 ──
BORDER = Border(
    left=Side(style='thin', color=C_BORDER),
    right=Side(style='thin', color=C_BORDER),
    top=Side(style='thin', color=C_BORDER),
    bottom=Side(style='thin', color=C_BORDER))
BORDER_NONE = Border()
BORDER_BOTTOM = Border(bottom=Side(style='medium', color=C_PRIMARY))

# ── 对齐 ──
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center')
ALIGN_WRAP   = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── 时段标准 ──
# 10个时段（含凌晨2+凌晨3合并）
REPORT_SHIFTS = ['凌晨1', '凌晨2+凌晨3', '早餐1', '早餐2', '午高峰', '下午茶1', '下午茶2', '晚高峰', '夜宵1', '夜宵2']
SHIFTS = [
    ('凌晨1', 0, 2, 1.5, 1),
    ('凌晨2+凌晨3', 2, 6, 1.3, 1),
    ('早餐1', 6, 8, 1.5, 1), ('早餐2', 8, 10.5, 2.0, 1),
    ('午高峰', 10.5, 13.5, 2.75, 6),
    ('下午茶1', 13.5, 15.5, 1.5, 2), ('下午茶2', 15.5, 17.5, 1.5, 2),
    ('晚高峰', 17.5, 20, 2.25, 5),
    ('夜宵1', 20, 22, 1.5, 2), ('夜宵2', 22, 24, 1.5, 2),
]
SN = [s[0] for s in SHIFTS]
SM = {s[0]: {'start': s[1], 'end': s[2], 'hours': s[3], 'orders': s[4]} for s in SHIFTS}

# ═══════════════════════════════════════
# 📐 辅助函数
# ═══════════════════════════════════════

def fmt_time(h):
    """小时→友好文字：2.75→'2小时45分钟', 2.25→'2小时15分钟', 0.2→'12分钟'"""
    if h is None or h == 0:
        return '0分钟'
    total_m = round(h * 60)
    hrs, mins = divmod(total_m, 60)
    if hrs == 0:
        return f'{mins}分钟'
    if mins == 0:
        return f'{hrs}小时'
    return f'{hrs}小时{mins}分钟'

def set_row(ws, row, height):
    ws.row_dimensions[row].height = height

def write_cell(ws, r, c, val, font=FONT_DATA, fill=FILL_BG, align=ALIGN_CENTER, border=BORDER):
    cell = ws.cell(r, c, val)
    cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = border

def write_header(ws, r, vals):
    set_row(ws, r, 32)
    for i, v in enumerate(vals, 1):
        write_cell(ws, r, i, v, FONT_HDR, FILL_HEADER)

def write_data(ws, r, vals, fill=FILL_BG, font=FONT_DATA):
    set_row(ws, r, 24)
    for i, v in enumerate(vals, 1):
        align = ALIGN_LEFT if i == 1 else ALIGN_CENTER
        write_cell(ws, r, i, v, font, fill, align)

def write_section(ws, r, text):
    set_row(ws, r, 30)
    for i in range(1, 12):
        c = ws.cell(r, i); c.fill = FILL_SECTION; c.border = BORDER_NONE
    ws.cell(r, 1).value = f'  {text}'
    ws.cell(r, 1).font = FONT_SECTION
    ws.cell(r, 1).alignment = ALIGN_LEFT

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def is_online(ws_val):
    """判断工作状态是否在线（新格式：1=忙碌, 3=小休, 5=离线）"""
    if ws_val is None or ws_val == '':
        return False
    # 处理字符串 '1'/'3' 或 int 1/3
    if isinstance(ws_val, str):
        v = ws_val.strip()
        if v.isdigit():
            return int(v) in (1, 3)
        return v in ('上班', '在线', '忙碌', '小休')
    if isinstance(ws_val, (int, float)):
        return int(ws_val) in (1, 3)
    return False


def format_status(ws_val):
    """格式化工作状态显示：上线/小休/下线"""
    if ws_val is None or ws_val == '':
        return '🔴 下线'
    if isinstance(ws_val, str):
        v = ws_val.strip()
        if v.isdigit():
            n = int(v)
            if n == 1: return '🟢 上线'
            if n == 3: return '🟡 小休'
            return '🔴 下线'
        if v in ('上班', '在线', '忙碌'): return '🟢 上线'
        if v == '小休': return '🟡 小休'
        return '🔴 下线'
    if isinstance(ws_val, (int, float)):
        n = int(ws_val)
        if n == 1: return '🟢 上线'
        if n == 3: return '🟡 小休'
        return '🔴 下线'
    return '🔴 下线'


def can_achieve_yn(text):
    """将estimate_achieve结果转为是/否"""
    if not text:
        return '否'
    if '已达成' in text or '可能' in text:
        return '是'
    return '否'

def auto_fit_cols(ws, max_col, skip_a=True, skip_row=1):
    """根据内容自动调整列宽（可跳过A列和标题行）"""
    for c in range(1, max_col + 1):
        if skip_a and c == 1: continue
        max_w = 6  # 最小宽度
        for r in range(1, ws.max_row + 1):
            if r == skip_row: continue
            v = ws.cell(r, c).value
            if v is not None:
                text = str(v)
                w = sum(2 if ord(ch) > 127 else 1 for ch in text)
                if w > max_w: max_w = w
        ws.column_dimensions[get_column_letter(c)].width = min(max_w + 2, 40)

# ═══════════════════════════════════════
# 🏗 主函数
# ═══════════════════════════════════════

def gen(fp=None):
    # ── 找文件（排除浏览器自动保存的空模板 < 100KB） ──
    def filter_real(fs):
        return [f for f in fs
                if not f.name.startswith('.') and not f.name.startswith('~')
                and f.stat().st_size > 10000]  # 空模板≈5KB, 真实数据>10KB
    if not fp:
        fls = sorted(Path.home().glob('Downloads/*出勤*.xlsx'),
                     key=lambda f: os.path.getmtime(f), reverse=True)
        fls = filter_real(fls)
        if not fls:
            fls = sorted(Path.home().glob('Downloads/2026*.xlsx'),
                        key=lambda f: os.path.getmtime(f), reverse=True)
            fls = filter_real(fls)
        if not fls:
            print('❌ 未找到出勤 Excel 文件'); return
    else:
        fls = [Path(fp)]
    fp = fls[0]
    print(f'📁 {Path(fp).name}')

    # ── 读取运单改派数据 ──
    combined_tf, transfer_only = lv_mod.read_transfer_data()
    if combined_tf:
        _total_transfer = sum(transfer_only.values())
        _total_all = sum(combined_tf.values())
        print(f'📋 改派扣减合计: {_total_transfer}单（内部全量扣减{_total_all}单）')
    else:
        _total_transfer = 0
        _total_all = 0

    # ── 运行分析 ──
    ra = att_mod.parse_attendance_format(str(fp))
    rl = lv_mod.analyze(str(fp), combined_tf, transfer_only)
    rf = fc_mod.analyze(str(fp), combined_tf, transfer_only)

    # ── 补充读取在线总时长（全天在线时长 vs 全天有效在线时长）──
    try:
        _rl_wb = __import__('openpyxl').load_workbook(str(fp), data_only=True)
        _rl_ws = _rl_wb.active
        _rl_headers = [_rl_ws.cell(1, c).value for c in range(1, _rl_ws.max_column + 1)]
        def _rl_find(targets):
            for t in targets:
                for i, h in enumerate(_rl_headers):
                    if h and t in str(h):
                        return i + 1
            return None
        _rl_ci = _rl_find(['全天在线时长'])
        _rl_seen = set()
        if _rl_ci:
            for _rl_r in range(2, _rl_ws.max_row + 1):
                _rl_nm = _rl_ws.cell(_rl_r, _rl_find(['姓名', '骑手姓名'])).value
                if not _rl_nm or _rl_nm in _rl_seen: continue
                _rl_seen.add(_rl_nm)
                _rl_raw = _rl_ws.cell(_rl_r, _rl_ci).value
                if _rl_nm in rl:
                    rl[_rl_nm]['online_total'] = lv_mod.parse_time(_rl_raw) if _rl_raw else rl[_rl_nm]['online']
        _rl_wb.close()
    except Exception:
        pass

    # ── 天气 & 尖峰场景（全局，供各Sheet使用） ──
    _weather_info = '未知'
    _weather_level = '正常'
    _is_weekend = datetime.now().weekday() >= 5
    _weekday_names = ['周一','周二','周三','周四','周五','周六','周日']
    _day_name = _weekday_names[datetime.now().weekday()]
    # 从运单文件读取天气等级数据
    try:
        _wf = sorted(Path.home().glob('Downloads/*运单*.xlsx'))
        _wf = [f for f in _wf if f.stat().st_size > 10000]
        if _wf:
            _wb = __import__('openpyxl').load_workbook(str(_wf[-1]), data_only=True)
            _ws = _wb.active
            _weather_map = {'正常':0,'轻微恶劣':1,'恶劣':2,'极恶劣':3,'罕见恶劣':4}
            _lv_counts = {0:0,1:0,2:0,3:0,4:0}
            _total = 0
            for _r in range(2, _ws.max_row + 1):
                _lv_raw = _ws.cell(_r, 40).value
                _w_raw = _ws.cell(_r, 41).value
                _total += 1
                if _lv_raw and str(_lv_raw).strip() in _weather_map:
                    _lv_counts[_weather_map[str(_lv_raw).strip()]] += 1
                if _r == 2 and _w_raw:
                    _weather_info = str(_w_raw).strip()
            _lv1234 = _lv_counts[1]+_lv_counts[2]+_lv_counts[3]+_lv_counts[4]
            _lv234 = _lv_counts[2]+_lv_counts[3]+_lv_counts[4]
            _lv34 = _lv_counts[3]+_lv_counts[4]
            if _lv1234/_total*100 <= 10:
                _weather_level = '正常天'
            elif _lv234/_total*100 <= 10:
                _weather_level = '轻微恶天'
            elif _lv34/_total*100 < 50:
                _weather_level = '一般恶天'
            else:
                _weather_level = '严重恶天'
    except Exception as _e:
        pass

    # ── 运力目标（全局，供标2-标5占比分母使用）+ 时段应排人数 ──
    # 从 aeolus 排班数据文件的排班统计 sheet 直读站点人数/应排人数/排班人数
    _sched_site_total = 0   # 站点人数
    _target_capacity = 0    # 应排人数
    _sched_count = 0        # 排班人数
    _period_targets = {}    # 各时段应排人数（排班目标）
    _period_scheduled = {}  # 各时段实排人数（排班人数，监控表口径）
    try:
        _fp_name = Path(fp).stem
        _date_match = re.search(r'(\d{4}-\d{2}-\d{2})', _fp_name)
        if _date_match:
            _sched_date = _date_match.group(1)
        else:
            _sched_date = _fp_name[:10]
        _sched_fls = sorted(Path.home().glob(f'Downloads/*{_sched_date}*排班数据*.xlsx'),
                           key=lambda f: os.path.getmtime(f), reverse=True)
        _sched_fls = [f for f in _sched_fls if not f.name.startswith('.') and not f.name.startswith('~')]
        if _sched_fls:
            _sched_wb = openpyxl.load_workbook(str(_sched_fls[0]), data_only=True)
            _sched_ws = None
            for _ssn in _sched_wb.sheetnames:
                if '排班统计' in _ssn:
                    _sched_ws = _sched_wb[_ssn]
                    break
            if _sched_ws:
                for _sr in range(2, _sched_ws.max_row + 1):
                    _period = str(_sched_ws.cell(_sr, 5).value or '')  # Col 5 = 时段名称（全天）
                    if _period == '全天':
                        _sched_site_total = int(_sched_ws.cell(_sr, 4).value or 0)   # Col 4 = 站点人数
                        _target_capacity = float(_sched_ws.cell(_sr, 6).value or 0)   # Col 6 = 应排人数
                        _sched_count = int(_sched_ws.cell(_sr, 7).value or 0)         # Col 7 = 排班人数
                    else:
                        # 存储各时段应排人数（排班目标），供达标率计算使用
                        _period_targets[_period] = float(_sched_ws.cell(_sr, 6).value or 0)
                        # 存储各时段实排人数（排班人数），与监控表口径对齐
                        _period_scheduled[_period] = int(_sched_ws.cell(_sr, 7).value or 0)
            _sched_wb.close()
    except Exception as _e:
        pass
    if _target_capacity == 0:
        _target_capacity = len([rv for rv in rl.values()]) if rl else 0  # fallback

    now = datetime.now()
    # 从数据源文件名提取日期和时间
    try:
        fp_name = Path(fp).stem
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fp_name)
        time_match = re.search(r'(\d{4}-\d{2}-\d{2})\s+(\d{2})_(\d{2})_(\d{2})', fp_name)
        if date_match:
            sd = date_match.group(1)
        else:
            sd = now.strftime('%Y-%m-%d')
        if time_match:
            ts = f'{time_match.group(2)}_{time_match.group(3)}_{time_match.group(4)}'
            # 用数据源导出时间作为分析基准时间
            dh = int(time_match.group(2)) + int(time_match.group(3)) / 60
        else:
            ts = now.strftime('%H_%M_%S')
            dh = now.hour + now.minute / 60
    except:
        sd = now.strftime('%Y-%m-%d')
        ts = now.strftime('%H_%M_%S')
        dh = now.hour + now.minute / 60
    
    data_hour = dh  # 用于后续班次/能否达成判断（基于数据导出时间）
    # 时段未达标过滤：历史数据（日期非今天）展示所有时段，当天数据按实际时间过滤
    _today_str = now.strftime('%Y-%m-%d')
    if sd and sd != _today_str:
        current_hour = 24  # 历史数据，展示所有时段
    else:
        current_hour = dh  # 当天数据，按实际时间过滤

    # 数据源时间戳（统一用于所有表头）
    _h, _m = ts.split('_')[0], ts.split('_')[1]
    _data_time = f'{_h}:{_m}'

    # 用数据源时间重新计算能否达成
    if rf:
        _lro = [18, 28, 35, 42, 53]
        _lrh = [6, 6, 6, 7, 8]
        for _key in ['all_candidates', 'lv3_candidates']:
            for _d in rf.get(_key, []):
                _eff = _d.get('effective_orders', _d.get('orders', 0))
                _on = _d.get('online', 0)
                _lv = _d.get('level', 0)
                _tgt_o = _lro[_lv] if _lv < 5 else 999
                _tgt_h = _lrh[_lv] if _lv < 5 else 999
                _rs = fc_mod.get_remaining_periods(_d.get('shift_schedule', ''), current_hour=data_hour)
                _d['remaining_shifts'] = '、'.join(_rs) if _rs else '无'
                _d['can_achieve'] = fc_mod.estimate_achieve(_eff, _on, _tgt_o, _rs, _d.get('delivering', 0), current_hour=data_hour, target_hours=_tgt_h if _lv < 5 else None)
        for _name, _d in rf.get('below_details', {}).items():
            _eff = _d.get('effective_orders', _d.get('orders', 0))
            _on = _d.get('online', 0)
            _rs = fc_mod.get_remaining_periods(_d.get('shift_schedule', ''), current_hour=data_hour)
            _d['remaining_shifts'] = '、'.join(_rs) if _rs else '无'
            _d['can_achieve'] = fc_mod.estimate_achieve(_eff, _on, 18, _rs, _d.get('delivering', 0), current_hour=data_hour, target_hours=6)

    wb = openpyxl.Workbook()

    # ── 时段参数 ──
    BASELINE_HOURS = 6
    BASELINE_ORDERS = 18

    # 10个时段权重（含凌晨2+凌晨3合并）
    PERIOD_WEIGHTS = {
        '凌晨1': 0.06, '凌晨2+凌晨3': 0.10,
        '早餐1': 0.06, '早餐2': 0.06,
        '午高峰': 0.20, '下午茶1': 0.06, '下午茶2': 0.06,
        '晚高峰': 0.20, '夜宵1': 0.10, '夜宵2': 0.10
    }

    # ── 从时段数据中合并凌晨2+凌晨3 ──
    period_data = {}
    if isinstance(ra, tuple) and len(ra) > 1 and ra[1]:
        # ra[1] 是站点名→periods 的映射
        for sn, sd_data in ra[1].items():
            pds = sd_data.get('periods', {})
            merged = {}
            for pn, pd_info in pds.items():
                merged[pn] = pd_info
            # 合并凌晨2+凌晨3
            if '凌晨2' in merged and '凌晨3' in merged:
                merged['凌晨2+凌晨3'] = {
                    'scheduled': merged['凌晨2']['scheduled'] + merged['凌晨3']['scheduled'],
                    'passed': merged['凌晨2']['passed'] + merged['凌晨3']['passed'],
                    'present': merged['凌晨2']['present'] + merged['凌晨3']['present'],
                }
            elif '凌晨2' in merged:
                merged['凌晨2+凌晨3'] = merged['凌晨2']
            # 写入各时段应排目标（排班计划值），监控表对齐用
            for _pn in list(merged.keys()):
                _pn_clean = _pn.replace('+', '、')
                _target = _period_targets.get(_pn, _period_targets.get(_pn_clean, None))
                if _target is not None:
                    merged[_pn]['target_capacity'] = _target
                elif '凌晨2' in _pn or '凌晨3' in _pn:
                    # 凌晨2+凌晨3合并时，排班统计无此行，设为0（与监控表一致）
                    merged[_pn]['target_capacity'] = 0
            period_data = merged
            # 用排班数据源的实排人数（Col7）覆盖实时考勤的计数，对齐监控表口径
            for _pn in list(period_data.keys()):
                _pn_clean = _pn.replace('+', '、')
                _sched_val = _period_scheduled.get(_pn, _period_scheduled.get(_pn_clean, None))
                if _sched_val is not None:
                    period_data[_pn]['scheduled'] = _sched_val
                elif '凌晨2' in _pn or '凌晨3' in _pn:
                    # 凌晨2+凌晨3合并时，排班统计无此行，设为0
                    if 'scheduled' in period_data[_pn]:
                        period_data[_pn]['scheduled'] = 0
            break  # 只取第一个站点

    # ═══════════════════════════════════════
    # Sheet 1: 全天达成
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = '全天达成'

    # 标题
    ws1.merge_cells('A1:J1')
    ws1['A1'].value = f'全天达成  {sd}  {_data_time}  v2.0'
    ws1['A1'].font = FONT_TITLE
    ws1['A1'].fill = FILL_HEADER
    ws1['A1'].alignment = ALIGN_LEFT
    for c in range(2, 10):
        ws1.cell(1, c).fill = FILL_HEADER
    set_row(ws1, 1, 40)

    # ── 尖峰场景加分/扣分行（第2行）──
    if rl:
        _s1_total_active = sum(1 for rv in rl.values() if rv['effective_orders'] > 0)
        _s1_level_counts = {i: 0 for i in range(6)}
        for rv in rl.values():
            if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS:
                _s1_level_counts[rv['level']] += 1
        _s1_b2b5 = _s1_level_counts.get(2,0)+_s1_level_counts.get(3,0)+_s1_level_counts.get(4,0)+_s1_level_counts.get(5,0)
        # 7月KPI新规：分母改为应排人数
        _s1_b2b5_pct = _s1_b2b5/_target_capacity*100 if _target_capacity else 0
        _s1_is_bad = _weather_level in ('轻微恶天','一般恶天','严重恶天')
        _s1_is_severe = _weather_level == '严重恶天'
        _s1_is_peak = _is_weekend or _s1_is_bad
        if not _s1_is_peak:
            _s1_fill = FILL_BG
        _s1_need_70 = int(_target_capacity * 0.70) + 1
        _s1_gap_70 = max(0, _s1_need_70 - _s1_b2b5)
        # 标2-标5占比 着色逻辑保留用于A2底色
        if _s1_b2b5_pct < 30:
            _s1_fill = FILL_FAIL
        elif _s1_b2b5_pct > 70:
            _s1_fill = FILL_PASS
        else:
            _s1_fill = FILL_WARN

        # 全天达成人数 & 排班总人数（提前算，A3要用）
        _s1_passed_cnt = sum(1 for rv in rl.values() if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS)
        _s1_scheduled_cnt = _sched_count if _sched_count > 0 else len(rl)
        _s1_day_attend_rate = _s1_passed_cnt / _s1_scheduled_cnt * 100 if _s1_scheduled_cnt > 0 else 0
        _s1_need_90 = int(_s1_scheduled_cnt * 0.90) if _s1_scheduled_cnt * 0.90 % 1 == 0 else int(_s1_scheduled_cnt * 0.90) + 1
        _s1_gap_90 = max(0, _s1_need_90 - _s1_passed_cnt)

        # 第2行：天气 & 星期
        _s1_scene_line = f'🌤 {_weather_info}/{_weather_level} | {_day_name}'
        ws1.cell(2, 1).value = _s1_scene_line
        ws1.cell(2, 1).font = Font(size=11, bold=True, color=C_TEXT, name='PingFang SC')
        ws1.cell(2, 1).fill = _s1_fill
        for _c in range(2, 10):
            ws1.cell(2, _c).fill = _s1_fill
        ws1.merge_cells('A2:J2')
        set_row(ws1, 2, 24)

        # 第3行：全天有效出勤率 & 缺口
        _s1_rate_str = f'{_s1_day_attend_rate:.1f}%'
        if _s1_gap_90 > 0:
            _s1_day_line = f'全天有效出勤率: {_s1_rate_str} | 差{_s1_gap_90}人达成90%及格线（需{_s1_need_90}人）'
        else:
            _s1_day_line = f'全天有效出勤率: {_s1_rate_str} ✅ 已达标'
        ws1.cell(3, 1).value = _s1_day_line
        ws1.cell(3, 1).font = Font(size=11, bold=True, color=C_TEXT, name='PingFang SC')
        # 达标绿色/未达标红色
        _s1_day_fill = FILL_PASS if _s1_day_attend_rate >= 90 else FILL_FAIL
        ws1.cell(3, 1).fill = _s1_day_fill
        for _c in range(2, 10):
            ws1.cell(3, _c).fill = _s1_day_fill
        ws1.merge_cells('A3:J3')
        set_row(ws1, 3, 24)

        # 第4行：改派扣减 + 尖峰标2-5
        _s1_jf_score = 2 if _s1_is_severe else 1
        if _s1_b2b5_pct < 30:
            _s1_jf_tip = f'标2-5占比{_s1_b2b5_pct:.1f}% ⚠️ 低于30% → 扣{_s1_jf_score}分'
        elif _s1_b2b5_pct > 70:
            _s1_jf_tip = f'标2-5占比{_s1_b2b5_pct:.1f}% ✅ 超过70% → 加{_s1_jf_score}分'
        else:
            _s1_jf_tip = f'标2-5占比{_s1_b2b5_pct:.1f}% 还差{_s1_gap_70}人到70%'
        _s1_transfer_line = f'📋 改派扣减合计: {_total_transfer}单 | {_s1_jf_tip}' if _total_transfer > 0 else f'📋 无扣减 | {_s1_jf_tip}'
        ws1.cell(4, 1).value = _s1_transfer_line
        ws1.cell(4, 1).font = Font(size=11, color=C_TEXT, name='PingFang SC')
        ws1.cell(4, 1).fill = _s1_fill
        for _c in range(2, 10):
            ws1.cell(4, _c).fill = _s1_fill
        ws1.merge_cells('A4:I4')
        set_row(ws1, 4, 24)

    if not rl:
        ws1.cell(3, 1).value = '⚠️ 无骑手数据'
        ws1.cell(2, 1).font = FONT_DATA
    else:
        # 区分达标/未达标
        # ── 找出非排班但活跃的骑手（不在rl中但原始数据有完单>0） ──
        _rl_names = set(rv['name'] for rv in rl.values())
        _unsched_active = []
        _wb_unsched = __import__('openpyxl').load_workbook(str(fp), data_only=True)
        _ws_unsched = _wb_unsched.active
        _unsched_seen = set()
        for _ur in range(2, _ws_unsched.max_row + 1):
            _uname = str(_ws_unsched.cell(_ur, 3).value or '').strip()
            if not _uname or _uname in _unsched_seen:
                continue
            _unsched_seen.add(_uname)
            if _uname in _rl_names:
                continue
            _uorders_raw = _ws_unsched.cell(_ur, 11).value
            try: _uorders = float(_uorders_raw or 0)
            except: _uorders = 0
            _uonline_raw = _ws_unsched.cell(_ur, 10).value
            _uonline_h = 0.0
            try:
                _um = __import__('re').search(r'(\d+)小时', str(_uonline_raw or ''))
                _umm = __import__('re').search(r'(\d+)分钟', str(_uonline_raw or ''))
                if _um: _uonline_h += float(_um.group(1))
                if _umm: _uonline_h += float(_umm.group(1)) / 60.0
            except: pass
            _udelivering_raw = _ws_unsched.cell(_ur, 13).value
            try: _udelivering = float(_udelivering_raw or 0)
            except: _udelivering = 0
            _uwork = str(_ws_unsched.cell(_ur, 8).value or '')
            _utf = transfer_only.get(_uname, 0)
            _ueff = max(0, _uorders - _utf)
            if _ueff > 0:
                _unsched_active.append({
                    'name': _uname, 'orders': int(_uorders),
                    'online': round(_uonline_h, 2),
                    'delivering': int(_udelivering),
                    'work_status': _uwork,
                    'shift_schedule': '',
                    'effective_orders': int(_ueff),
                })
        _wb_unsched.close()

        # 合并排班+非排班骑手到统一列表
        _all_riders = list(rl.values()) + _unsched_active
        passed_riders = [rv for rv in _all_riders if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS]
        failed_riders = [rv for rv in _all_riders if not (rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS)]

        # 从排班数据文件获取站点人数/排班人数（已在上方 _sched_site_total / _sched_count 中读取）
        total = _sched_site_total if _sched_site_total > 0 else (int(_target_capacity) if _target_capacity > 0 else len(rl))
        scheduled = _sched_count if _sched_count > 0 else len(rl)
        passed = len([rv for rv in rl.values() if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS])
        failed = len([rv for rv in rl.values() if not (rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS)])

        # 统计信息行（第5行）
        # 活跃骑手：监控表口径=全部骑手中全天有效完成单>0的人数
        # 直接从出勤数据读取全部96名骑手（含非排班），统计完单量>0
        _all_active = set()
        _all_riders_orders = {}  # rider_id -> (name, max_orders)
        _wb_all = __import__('openpyxl').load_workbook(str(fp), data_only=True)
        _ws_all = _wb_all.active
        for _r in range(2, _ws_all.max_row + 1):
            _rid = str(_ws_all.cell(_r, 4).value or '')
            _nm = str(_ws_all.cell(_r, 3).value or '')
            _od_raw = _ws_all.cell(_r, 11).value
            try: _od = float(_od_raw or 0)
            except: _od = 0
            if _rid and _rid not in _all_riders_orders:
                _all_riders_orders[_rid] = (_nm, _od)
            elif _rid and _rid in _all_riders_orders:
                _existing_od = _all_riders_orders[_rid][1]
                if _od > _existing_od:
                    _all_riders_orders[_rid] = (_nm, _od)
        _wb_all.close()
        for _rid, (_nm, _od) in _all_riders_orders.items():
            _tf = transfer_only.get(_nm, 0)
            _eff = max(0, _od - _tf)
            if _eff > 0:
                _all_active.add(_rid)
        active_riders = len(_all_active)
        _unsched_cnt = len([d for d in _unsched_active if d['orders'] > 0])
        _unsched_passed = [d for d in _unsched_active if d['orders'] >= BASELINE_ORDERS and d['online'] >= BASELINE_HOURS]
        _sched_passed = len(passed_riders) - len(_unsched_passed)
        info_line1 = f'站点人数: {total}人  当天排班人数: {scheduled}人  应排人数: {int(_target_capacity)}人  活跃骑手: {active_riders}人'
        info_line2 = f'达成人数: {len(passed_riders)}人  未达成人数: {len(failed_riders)}人  排班达标: {_sched_passed}人'
        if _unsched_passed:
            info_line2 += f'  非排班达标: {len(_unsched_passed)}人'
        ws1.cell(5, 1).value = info_line1
        ws1.cell(5, 1).font = FONT_MUTED
        ws1.cell(5, 1).alignment = ALIGN_LEFT
        set_row(ws1, 5, 22)
        ws1.cell(6, 1).value = info_line2
        ws1.cell(6, 1).font = FONT_MUTED
        ws1.cell(6, 1).alignment = ALIGN_LEFT
        set_row(ws1, 6, 22)
        r = 7

        # ── 未达成明细 ──
        _show_failed = len(failed_riders)
        write_section(ws1, r, f'❌ 未达成（≥{BASELINE_HOURS}h / ≥{BASELINE_ORDERS}单）  {_show_failed}人'); r += 1

        if failed_riders:
            headers = ['排名', '骑手', '完单量', '有效时长', '配送中', '是否在线', '后续班次', '能否达成', '差单', '差时']
            write_header(ws1, r, headers); r += 1

            for i, rv in enumerate(sorted(failed_riders,
                                          key=lambda x: (-x['orders'], -x['online'], x['name']))):
                orders = rv['orders']
                online = rv['online']
                delivering = rv.get('delivering', 0)

                ot = '🟢 在线' if is_online(rv.get('work_status', '')) else '🔴 离线'

                go_int = max(0, BASELINE_ORDERS - orders)
                gh_float = max(0, round(BASELINE_HOURS - online, 1))

                # 后续班次 + 能否达成（使用全天数据中的排班时段）
                _sched = rv.get('shift_schedule', '')
                _rs = fc_mod.get_remaining_periods(_sched, current_hour=data_hour) if _sched else []
                _rs_str = '、'.join(_rs) if _rs else '无'
                _eff = rv.get('effective_orders', orders)
                _ca = fc_mod.estimate_achieve(_eff, online, 18, _rs, delivering, current_hour=data_hour, target_hours=6)
                _ca_yn = can_achieve_yn(_ca)

                _is_unsched = rv['name'] not in _rl_names
                fill = FILL_UNSCHED if _is_unsched else (FILL_FAIL if i % 2 == 0 else FILL_BG)
                write_data(ws1, r, [i + 1, rv['name'],
                                     f'{orders:.0f}单', fmt_time(online),
                                     f'{delivering:.0f}单' if delivering > 0 else '无',
                                     ot, _rs_str, _ca_yn,
                                     f'{go_int:.0f}单' if go_int > 0 else '—',
                                     fmt_time(gh_float) if gh_float > 0 else '—'],
                           fill=fill); r += 1

            # 能否达成着色
            first_ca_row = r - len(failed_riders)
            for j in range(len(failed_riders)):
                _cr = first_ca_row + j
                cv = str(ws1.cell(_cr, 8).value or '')
                if cv == '是':
                    ws1.cell(_cr, 8).font = Font(size=10, color='38A169', name='PingFang SC')
                elif cv == '否':
                    ws1.cell(_cr, 8).font = Font(size=10, color='E53E3E', name='PingFang SC')
                elif '困难' in cv or '可能' in cv:
                    ws1.cell(_cr, 8).font = Font(size=10, color='D69E2E', name='PingFang SC')

        auto_fit_cols(ws1, 10)

    # ═══════════════════════════════════════
    # Sheet 2: 时段未达标
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet('时段未达标')

    # 标题
    ws2.merge_cells('A1:K1')
    ws2['A1'].value = f'时段日报  {sd}  {_data_time}'
    ws2['A1'].font = FONT_TITLE
    ws2['A1'].fill = FILL_HEADER
    ws2['A1'].alignment = ALIGN_LEFT
    for c in range(2, 12):
        ws2.cell(1, c).fill = FILL_HEADER
    set_row(ws2, 1, 40)

    # 全天时段达标率（对齐监控表横向格式）
    ws2.merge_cells('A2:F2')
    ws2.cell(2, 1).value = '全天时段达标率'
    ws2.cell(2, 1).font = FONT_MUTED
    ws2.cell(2, 1).alignment = ALIGN_CENTER
    set_row(ws2, 2, 22)
    ws2.column_dimensions['A'].width = 14
    for _c in ['B','C','D','E','F']:
        ws2.column_dimensions[_c].width = 13

    period_order = ['凌晨1', '凌晨2+凌晨3', '早餐1', '早餐2', '午高峰',
                    '下午茶1', '下午茶2', '晚高峰', '夜宵1', '夜宵2']

    PERIOD_HOURS = {
        '凌晨1': (0, 2), '凌晨2+凌晨3': (2, 6),
        '早餐1': (6, 8), '早餐2': (8, 10.5),
        '午高峰': (10.5, 13.5), '下午茶1': (13.5, 15.5), '下午茶2': (15.5, 17.5),
        '晚高峰': (17.5, 20), '夜宵1': (20, 22), '夜宵2': (22, 24),
    }

    # 收集各时段数据（含未开始的时段，监控表全展示）
    period_vals = {}
    aggr_sum = 0.0
    aggr_w = 0.0

    for pn in period_order:
        pd_info = period_data.get(pn, {})
        s_val = pd_info.get('scheduled', 0) if isinstance(pd_info, dict) else 0
        tgt_val = pd_info.get('target_capacity', s_val) if isinstance(pd_info, dict) else s_val
        p_val = pd_info.get('passed', 0) if isinstance(pd_info, dict) else 0

        # 达标率 = passed / target_capacity，与监控表对齐
        pass_rate = p_val / tgt_val if tgt_val > 0 else 1.0

        w = PERIOD_WEIGHTS.get(pn, 0)
        aggr_sum += min(pass_rate, 1.0) * w
        aggr_w += w

        period_vals[pn] = {
            'tgt': int(tgt_val),
            'sched': int(s_val),
            'passed': int(p_val),
            'rate': pass_rate,
        }

    # Row 3: 加权汇总（百分比2位小数，对齐监控表格式）
    if aggr_w > 0:
        ws2.merge_cells('A3:F3')
        ws2.cell(3, 1).value = aggr_sum / aggr_w
        ws2.cell(3, 1).font = FONT_BOLD
        ws2.cell(3, 1).alignment = ALIGN_CENTER
        ws2.cell(3, 1).number_format = '0.00%'
        set_row(ws2, 3, 22)

    r2 = 4

    def _write_block(ws, row, pns):
        """写一个监控表横排块：项目 / 应排目标 / 实排目标 / 达标骑手 / 达标率（含框线）"""
        labels = ['项目', '应排目标', '实排目标', '达标骑手', '达标率']
        keys = [None, 'tgt', 'sched', 'passed', 'rate']
        for li, (lb, kk) in enumerate(zip(labels, keys)):
            r = row + li
            if kk is None:
                # 表头行（项目/时段名）→ 深蓝底 + 白字 + 框线
                ws.cell(r, 1).value = lb
                ws.cell(r, 1).font = FONT_HDR
                ws.cell(r, 1).fill = FILL_HEADER
                ws.cell(r, 1).alignment = ALIGN_CENTER
                ws.cell(r, 1).border = BORDER
                for ci, pn in enumerate(pns):
                    ws.cell(r, 2+ci).value = pn
                    ws.cell(r, 2+ci).font = FONT_HDR
                    ws.cell(r, 2+ci).fill = FILL_HEADER
                    ws.cell(r, 2+ci).alignment = ALIGN_CENTER
                    ws.cell(r, 2+ci).border = BORDER
            else:
                # 数据行（应排/实排/达标/达标率）→ 标签列浅底，数据列普通，全部加框线
                ws.cell(r, 1).value = lb
                ws.cell(r, 1).font = Font(size=10, bold=True, color=C_TEXT, name='PingFang SC')
                ws.cell(r, 1).fill = FILL_BG
                ws.cell(r, 1).alignment = ALIGN_CENTER
                ws.cell(r, 1).border = BORDER
                for ci, pn in enumerate(pns):
                    if kk == 'rate':
                        rate = period_vals.get(pn, {}).get('rate', 1.0)
                        ws.cell(r, 2+ci).value = round(min(rate, 1.0), 1)
                        ws.cell(r, 2+ci).number_format = '0.0'
                        ws.cell(r, 2+ci).font = FONT_PASS
                    else:
                        ws.cell(r, 2+ci).value = period_vals.get(pn, {}).get(kk, 0)
                        ws.cell(r, 2+ci).font = FONT_DATA
                    ws.cell(r, 2+ci).fill = FILL_BG
                    ws.cell(r, 2+ci).alignment = ALIGN_CENTER
                    ws.cell(r, 2+ci).border = BORDER
            set_row(ws, r, 22)
        return row + len(labels)

    # Block 1: 凌晨1 ~ 午高峰
    r2 = _write_block(ws2, r2, ['凌晨1', '凌晨2+凌晨3', '早餐1', '早餐2', '午高峰'])
    # Block 2: 下午茶1 ~ 夜宵2
    r2 = _write_block(ws2, r2, ['下午茶1', '下午茶2', '晚高峰', '夜宵1', '夜宵2'])

    # ── 总达标率（用于后续考核判定）──
    _total_rate = aggr_sum / aggr_w * 100 if aggr_w > 0 else 0

    # ── 门槛/封顶判定（考核方案）──
    _baseline_achieved = sum(1 for rv in rl.values() if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS) if rl else 0
    _scheduled_total = sum(1 for rv in rl.values()) if rl else 0  # 排班总人数
    _threshold = int(_scheduled_total * 0.85)  # 门槛值
    _cap = int(_scheduled_total * 1.10)  # 封顶值
    r2 += 1
    ws2.merge_cells(f'A{r2}:F{r2}')
    _pass_rate_check = _total_rate >= 85
    _pass_achieve_check = _baseline_achieved >= _threshold
    _judge_parts = []
    if _pass_rate_check:
        _judge_parts.append(f'✅ 时段达标率{_total_rate:.1f}% ≥ 85%')
    else:
        _judge_parts.append(f'❌ 时段达标率{_total_rate:.1f}% ＜ 85%')
    if _pass_achieve_check:
        _judge_parts.append(f'✅ 达标{_baseline_achieved}人 ≥ 门槛{_threshold}人')
    else:
        _judge_parts.append(f'❌ 达标{_baseline_achieved}人 ＜ 门槛{_threshold}人')
    if _baseline_achieved > _cap:
        _judge_parts.append(f'⚠️ 超封顶{_cap}人，超出{_baseline_achieved-_cap}人不计费')
    else:
        _judge_parts.append(f'封顶{_cap}人（未超）')
    _result = '✅ 达标，可结算基础服务费' if (_pass_rate_check and _pass_achieve_check) else '❌ 不达标，不结算基础服务费'
    _judge_text = f'🏁 考核判定: {" | ".join(_judge_parts)} | {_result}'
    _judge_fill = FILL_PASS if (_pass_rate_check and _pass_achieve_check) else FILL_FAIL
    ws2.cell(r2, 1).value = _judge_text
    ws2.cell(r2, 1).font = Font(size=10, bold=True, color=C_TEXT, name='PingFang SC')
    ws2.cell(r2, 1).fill = _judge_fill
    ws2.cell(r2, 1).border = BORDER
    for _c in range(2, 7):
        ws2.cell(r2, _c).fill = _judge_fill
        ws2.cell(r2, _c).border = BORDER
    ws2.cell(r2, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    set_row(ws2, r2, 40)

    # ── 构建骑手最后班次查询 ──
    rider_last_shift = {}
    for _rn, _rd in rl.items():
        _sched = _rd.get('shift_schedule', '')
        # 适配逗号、顿号、空格等分隔符
        _parts = [s.strip() for s in re.split(r'[,、\s]+', _sched) if s.strip()]
        rider_last_shift[_rn] = _parts[-1] if _parts else '—'
    # 补充低于基线骑手的班次信息
    for _rn, _rd in rf.get('below_details', {}).items():
        if _rn not in rider_last_shift:
            _sched = _rd.get('shift_schedule', '')
            _parts = [s.strip() for s in re.split(r'[,、\s]+', _sched) if s.strip()]
            rider_last_shift[_rn] = _parts[-1] if _parts else '—'

    # ── 各时段未达标明细 ──
    r2 += 1
    write_section(ws2, r2, '各时段未达标明细')
    r2 += 1

    has_fail_detail = False
    for pn in period_order:
        ps, pe = PERIOD_HOURS.get(pn, (0, 24))
        # 跳过未开始的时段
        if ps >= current_hour:
            continue

        pd_info = period_data.get(pn, {})
        s_val = pd_info.get('scheduled', 0) if isinstance(pd_info, dict) else 0
        if s_val == 0:
            continue

        # 获取该时段未达标骑手
        fails = []
        unsched_passed = []  # 非排班但达标的
        if isinstance(ra, tuple) and ra[0]:
            fails = [d for d in ra[0]
                     if d.get('period') == pn and d.get('scheduled') and not d.get('passed')]
            unsched_passed = [d for d in ra[0]
                              if d.get('period') == pn and not d.get('scheduled') and d.get('passed')]

        if not fails and not unsched_passed:
            continue

        # 判断时段是否正在进行中
        is_ongoing = ps <= current_hour < pe

        has_fail_detail = True
        write_section(ws2, r2, f'  {pn}')
        r2 += 1

        if is_ongoing:
            # 正在进行的时段：加预测列 + 剩余时间（不显示判定）
            headers3 = ['骑手', '实完单', '有效时长', '要求单', '要求时', '差单', '差时', '剩余时间', '工作状态', '最后班次', '配送中', '预测']
            write_header(ws2, r2, headers3)
        else:
            headers3 = ['骑手', '实完单', '有效时长', '要求单', '要求时', '差单', '差时', '最后班次', '判定']
            write_header(ws2, r2, headers3)
        r2 += 1

        # 排班未达标
        for d in sorted(fails, key=lambda x: x.get('name', '')):
            ao = int(d.get('orders', 0))
            ro = int(d.get('req_orders', 0))
            ah = d.get('online', 0)
            rh = d.get('req_hours', 0)

            go = max(0, ro - ao)
            gh = max(0, round(rh - ah, 2))

            g6 = '✓' if go <= 0 else f'{go:.0f}'
            g7 = '✓' if gh <= 0 else f'-{fmt_time(abs(gh))}'

            if go > 0 and gh > 0:
                jd = '双不达标'
            elif go > 0:
                jd = '完单不达标'
            elif gh > 0:
                jd = '时长不达标'
            else:
                jd = '达标'

            remaining_h = 0
            if is_ongoing:
                # 预测逻辑
                work = d.get('work_status', '')
                delivering = d.get('delivering', 0)
                remaining_h = max(0, pe - current_hour)

                # 只要时段还有剩余时间，不判死
                base_orders = ao
                base_time = ah
                predicted_time = base_time + remaining_h
                t_ok = predicted_time >= rh

                # 单是系统派的，只要时间够+上班中就行
                # 只有剩余时间<30min、已下班、或申请小休时才看单量
                if remaining_h >= 0.5 and work in ('上班', '上线', '小休'):
                    # 时间够+在上班，系统会继续派单
                    prediction = '是' if t_ok else '否'
                else:
                    # 时间紧迫或已下班，看配送中单能否补齐
                    predicted_orders = base_orders + delivering
                    o_ok = predicted_orders >= ro
                    prediction = '是' if o_ok and t_ok else '否'
                remaining_str = fmt_time(remaining_h)

                _last_shift = rider_last_shift.get(d.get('name', ''), '—')
                write_data(ws2, r2, [d.get('name', ''), f'{ao:.0f}', fmt_time(ah),
                                       f'{ro:.0f}', fmt_time(rh), g6, g7, remaining_str, work, _last_shift, f'{delivering:.0f}', prediction])
            else:
                _last_shift = rider_last_shift.get(d.get('name', ''), '—')
                write_data(ws2, r2, [d.get('name', ''), f'{ao:.0f}', fmt_time(ah),
                                       f'{ro:.0f}', fmt_time(rh), g6, g7, _last_shift, jd])

            # 差单/差时着色
            g6_val = str(ws2.cell(r2, 6).value or '')
            g7_val = str(ws2.cell(r2, 7).value or '')
            if g6_val == '✓':
                ws2.cell(r2, 6).fill = FILL_PASS
                ws2.cell(r2, 6).font = FONT_PASS
            elif g6_val:
                ws2.cell(r2, 6).fill = FILL_FAIL
                ws2.cell(r2, 6).font = FONT_FAIL

            if g7_val == '✓':
                ws2.cell(r2, 7).fill = FILL_PASS
                ws2.cell(r2, 7).font = FONT_PASS
            elif g7_val and not (is_ongoing and gh > 0 and gh < remaining_h):
                ws2.cell(r2, 7).fill = FILL_FAIL
                ws2.cell(r2, 7).font = FONT_FAIL

            # 预测列着色：否=标红，是=标绿
            if is_ongoing:
                _pred_col = 12  # 预测列在第12列
                _pred_val = str(ws2.cell(r2, _pred_col).value or '')
                if _pred_val == '否':
                    ws2.cell(r2, _pred_col).fill = FILL_FAIL
                    ws2.cell(r2, _pred_col).font = FONT_FAIL
                elif _pred_val == '是':
                    ws2.cell(r2, _pred_col).fill = FILL_PASS
                    ws2.cell(r2, _pred_col).font = FONT_PASS

            r2 += 1

        # 非排班但达成
        if unsched_passed:
            _usc_hdr = 12 if is_ongoing else 9  # 表头列数
            _usc_dat = 9  # 非排班数据列数（无预测列）
            ws2.cell(r2, 1).value = '  -- 非排班达成 --'
            ws2.cell(r2, 1).font = Font(size=10, color=C_BLUE, bold=True, name='PingFang SC')
            ws2.cell(r2, 1).alignment = ALIGN_LEFT
            for _c in range(1, _usc_hdr+1):
                ws2.cell(r2, _c).fill = FILL_UNSCHED
                ws2.cell(r2, _c).border = BORDER
            for _c in range(1, _usc_hdr+1):
                ws2.cell(r2, _c).border = BORDER
            set_row(ws2, r2, 20)
            r2 += 1

            for d in sorted(unsched_passed, key=lambda x: x.get('name', '')):
                ao = int(d.get('orders', 0))
                ro = int(d.get('req_orders', 0))
                ah = d.get('online', 0)
                rh = d.get('req_hours', 0)

                go = max(0, ro - ao)
                gh = max(0, round(rh - ah, 2))

                g6 = '✓' if go <= 0 else f'{go:.0f}'
                g7 = '✓' if gh <= 0 else f'-{fmt_time(abs(gh))}'

                jd = '双不达标'
                if go > 0 and gh > 0:
                    jd = '双不达标'
                elif go > 0:
                    jd = '完单不达标'
                elif gh > 0:
                    jd = '时长不达标'
                else:
                    jd = '✅ 达成'

                _last_shift = rider_last_shift.get(d.get('name', ''), '—')
                write_data(ws2, r2, [d.get('name', ''), f'{ao:.0f}', fmt_time(ah),
                                       f'{ro:.0f}', fmt_time(rh), g6, g7, _last_shift, jd])

                # 整行浅蓝底色
                for _c in range(1, _usc_hdr+1):
                    ws2.cell(r2, _c).fill = FILL_UNSCHED

                # 差单/差时着色
                g6_val = str(ws2.cell(r2, 6).value or '')
                g7_val = str(ws2.cell(r2, 7).value or '')
                if g6_val == '✓':
                    ws2.cell(r2, 6).fill = FILL_PASS
                    ws2.cell(r2, 6).font = FONT_PASS
                if g7_val == '✓':
                    ws2.cell(r2, 7).fill = FILL_PASS
                    ws2.cell(r2, 7).font = FONT_PASS

                r2 += 1

    if not has_fail_detail:
        ws2.cell(r2, 1).value = '  所有时段均已达标 ✅'
        ws2.cell(r2, 1).font = FONT_PASS

    auto_fit_cols(ws2, 8)

    # ═══════════════════════════════════════
    # Sheet 3: 档位达成
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet('档位达成')

    ws3.merge_cells('A1:D1')
    ws3['A1'].value = f'档位达成  {sd}  {_data_time}'
    ws3['A1'].font = FONT_TITLE
    ws3['A1'].fill = FILL_HEADER
    ws3['A1'].alignment = ALIGN_LEFT
    for c in range(2, 5):
        ws3.cell(1, c).fill = FILL_HEADER
    set_row(ws3, 1, 40)
    ws3.column_dimensions['A'].width = 14

    if rl:
        passed_riders = [rv for rv in rl.values() if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS]
        level_3_plus = sum(1 for r in passed_riders if r['level'] >= 3)
        level_2_to_5 = sum(1 for r in passed_riders if 2 <= r['level'] <= 5)
        potential = level_3_plus + len(rf.get('lv3_candidates', [])) if rf else level_3_plus

        ws3.cell(2, 1).value = f'{len(passed_riders)}人    二档至五档{level_2_to_5}人（{level_2_to_5 / len(passed_riders) * 100:.1f}%）    潜力{potential}人'
        ws3.cell(2, 1).font = FONT_MUTED
        ws3.cell(2, 1).alignment = ALIGN_LEFT
        set_row(ws3, 2, 22)

        r3 = 5
        write_header(ws3, r3, ['档位', '标准', '人数', '占比'])
        r3 += 1

        level_counts = {i: 0 for i in range(6)}
        for rv in passed_riders:
            level_counts[rv['level']] += 1

        level_emojis = {5: '🏆 ', 4: '💪 ', 3: '⭐ ', 2: '', 1: '', 0: ''}
        level_names = {5: '标准5', 4: '标准4', 3: '标准3', 2: '标准2', 1: '标准1', 0: '不足标准1'}
        level_stds = {5: '8h/53单', 4: '7h/42单', 3: '6h/35单', 2: '6h/28单', 1: '6h/18单', 0: '—'}
        level_keys = [5, 4, 3, 2, 1, 0]

        _b2b5_cnt = level_counts.get(2,0) + level_counts.get(3,0) + level_counts.get(4,0) + level_counts.get(5,0)
        _b2b5_pct = _b2b5_cnt / _target_capacity * 100 if _target_capacity else 0
        _is_bad_weather = _weather_level in ('轻微恶天','一般恶天','严重恶天')
        _is_severe = _weather_level == '严重恶天'
        _is_peak_scene = _is_weekend or _is_bad_weather
        _jf_score = 2 if _is_severe else 1
        if _is_peak_scene:
            _jf_label = f'尖峰场景（{"周末 " if _is_weekend else ""}{"严重恶天" if _is_severe else "恶天" if _is_bad_weather else ""}）'
        else:
            _jf_label = '非尖峰场景'
        _need_70 = int(_target_capacity * 0.70) + 1
        _gap_70 = max(0, _need_70 - _b2b5_cnt)
        if not _is_peak_scene:
            _jf_result = '非尖峰场景，不考核'
        elif _b2b5_pct < 30:
            _jf_result = f'⚠️ 低于30% | 当天扣{_jf_score}分'
        elif _b2b5_pct > 70:
            _jf_result = f'✅ 超过70% | 当天加{_jf_score}分'
        else:
            _jf_result = f'还差{_gap_70}人达标2→加{_jf_score}分'
        # ── 天气 + 尖峰场景 ──
        ws3.cell(3, 1).value = f'🌤 {_weather_info}/{_weather_level} | {_day_name} | {_jf_label}'
        ws3.cell(3, 1).font = FONT_MUTED
        ws3.cell(3, 1).alignment = ALIGN_LEFT
        set_row(ws3, 3, 22)

        # ── 标2-5占比评分 + 改派单量 ──
        _s3_b2b5_fill = FILL_PASS if _b2b5_pct > 70 else (FILL_FAIL if _b2b5_pct < 30 else FILL_BG)
        if not _is_peak_scene:
            _s3_b2b5_text = f'标2-5占比: {_b2b5_pct:.1f}% | 非尖峰场景，不考核'
        elif _b2b5_pct < 30:
            _s3_b2b5_text = f'标2-5占比: {_b2b5_pct:.1f}% ⚠️ 低 → 扣{_jf_score}分'
        elif _b2b5_pct > 70:
            _s3_b2b5_text = f'标2-5占比: {_b2b5_pct:.1f}% ✅ 高 → 加{_jf_score}分'
        else:
            _s3_b2b5_text = f'标2-5占比: {_b2b5_pct:.1f}% → 不加不减'
        ws3.cell(4, 1).value = f'{_s3_b2b5_text}   |   📋 改派: {_total_transfer}单'
        ws3.cell(4, 1).font = Font(size=10, bold=True, color=C_TEXT, name='PingFang SC')
        ws3.cell(4, 1).alignment = ALIGN_LEFT
        ws3.cell(4, 1).fill = _s3_b2b5_fill
        set_row(ws3, 4, 22)

        for lv in level_keys:
            cnt = level_counts.get(lv, 0)
            pct = f'{cnt / len(passed_riders) * 100:.1f}%' if passed_riders else '0%'
            write_data(ws3, r3, [f'{level_emojis[lv]}{level_names[lv]}', level_stds[lv], cnt, pct])
            if lv >= 4: ws3.cell(r3,1).font = FONT_BOLD
            r3 += 1

        # ── 接近升级骑手 ──
        r3 += 1
        write_section(ws3, r3, '接近升级  ·  差1单或30分钟内')
        r3 += 1

        write_header(ws3, r3, ['骑手', '当前', '实完单', '有效时长', '差距'])
        r3 += 1

        next_orders = [18, 28, 35, 42, 53]
        next_hours = [6, 6, 6, 7, 8]
        for rv in sorted(passed_riders, key=lambda x: (-x['level'], x['name'])):
            lv = rv['level']
            if lv >= 5: continue

            eff_orders = rv.get('effective_orders', rv['orders'])
            go = max(0, next_orders[lv] - eff_orders)
            gh = max(0, round(next_hours[lv] - rv['online'], 1))
            if go > 1: continue

            gaps = []
            if go > 0: gaps.append(f'差{go:.0f}单')
            if gh > 0: gaps.append(f'差{fmt_time(gh)}')
            write_data(ws3, r3, [rv['name'], level_names[lv],
                                   f'{eff_orders:.0f}单', fmt_time(rv['online']),
                                   '、'.join(gaps) if gaps else ''])
            r3 += 1

        # ── 低于基线骑手 ──
        failed_riders = [rv for rv in rl.values() if not (rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS)]
        if failed_riders:
            r3 += 1
            write_section(ws3, r3, f'低于基线（<{BASELINE_HOURS}h 或 <{BASELINE_ORDERS}单）  {len(failed_riders)}人')
            r3 += 1

            write_header(ws3, r3, ['骑手', '完单量', '有效时长', '差距'])
            r3 += 1

            for rv in sorted(failed_riders, key=lambda x: (-x['orders'], -x['online'], x['name'])):
                go = max(0, BASELINE_ORDERS - rv['orders'])
                gh = max(0, round(BASELINE_HOURS - rv['online'], 1))
                gaps = []
                if go > 0: gaps.append(f'差{go:.0f}单')
                if gh > 0: gaps.append(f'差{fmt_time(gh)}')
                fill_bl = FILL_ALT if r3 % 2 == 0 else FILL_BG
                write_data(ws3, r3, [rv['name'], f'{rv["orders"]:.0f}单',
                                       fmt_time(rv['online']),
                                       '、'.join(gaps) if gaps else ''],
                                       fill=fill_bl)
                r3 += 1

        auto_fit_cols(ws3, 5)

    # ═══════════════════════════════════════
    # Sheet 4: 重点跟进
    # ═══════════════════════════════════════
    ws4 = wb.create_sheet('重点跟进')

    ws4.merge_cells('A1:I1')
    ws4['A1'].value = f'重点跟进  {sd}  {_data_time}'
    ws4['A1'].font = FONT_TITLE
    ws4['A1'].fill = FILL_HEADER
    ws4['A1'].alignment = ALIGN_LEFT
    for c in range(2, 9):
        ws4.cell(1, c).fill = FILL_HEADER
    set_row(ws4, 1, 40)

    for cl in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws4.column_dimensions[cl].width = 16

    if rf:
        lv3_candidates = rf.get('lv3_candidates', [])
        all_candidates = rf.get('all_candidates', [])

        # 分组
        groups = {
            '📈 冲标准5  标准4→标准5（8h/53单）': [],
            '📈 冲标准4  标准3→标准4（7h/42单）': [],
            '🎯 冲标准3  标准2→标准3（6h/35单）': [],
            '📈 冲标准2  标准1→标准2（6h/28单）': [],
            '📈 冲标准1  不足标准1→标准1（6h/18单）': [],
        }
        target_map = {
            '标准5': '📈 冲标准5  标准4→标准5（8h/53单）',
            '标准4': '📈 冲标准4  标准3→标准4（7h/42单）',
            '标准3': '🎯 冲标准3  标准2→标准3（6h/35单）',
            '标准2': '📈 冲标准2  标准1→标准2（6h/28单）',
            '标准1': '📈 冲标准1  不足标准1→标准1（6h/18单）',
        }

        for d in all_candidates:
            gk = target_map.get(d.get('target', ''))
            if gk:
                groups[gk].append(d)
        # 低于基线的冲标准1候选
        for _bn, _bd in rf.get('below_details', {}).items():
            _go = _bd.get('gap_orders', 0)
            _gh = _bd.get('gap_hours', 0)
            if (0 < _go <= 2) or (0 < _gh <= 0.33):
                groups['📈 冲标准1  不足标准1→标准1（6h/18单）'].append({
                    'name': _bn, 'level': 0, 'target': '标准1',
                    'gap_o': _go, 'gap_h': _gh,
                    'effective_orders': _bd.get('effective_orders', 0),
                    'orders': _bd.get('orders', 0),
                    'online': _bd.get('online', 0),
                    'delivering': _bd.get('delivering', 0),
                    'transfer_orders': _bd.get('transfer_orders', 0),
                    'work_status': _bd.get('work_status', ''),
                    'remaining_shifts': _bd.get('remaining_shifts', ''),
                    'can_achieve': _bd.get('can_achieve', ''),
                })

        r4 = 3
        total_up = 0

        for st, items in groups.items():
            if not items: continue
            items.sort(key=lambda x: x.get('gap_o', 2) * 2 + x.get('gap_h', 0.33) * 6)
            total_up += len(items)

            r4 += 1
            write_section(ws4, r4, f'{st}  —  {len(items)}人')
            r4 += 1

            headers4 = ['骑手', '当前档位', '在线状态', '配送中单量', '实完单', '有效时长', '差距', '后续班次', '能否达成']
            write_header(ws4, r4, headers4)
            r4 += 1

            for d in items:
                go = d.get('gap_o', 0)
                gh = d.get('gap_h', 0)
                ot = '🟢 在线' if is_online(d.get('work_status', '')) else '🔴 离线'
                gp = []
                if go > 0: gp.append(f'差{go:.0f}单')
                if gh > 0: gp.append(f'差{fmt_time(gh)}')
                lv = d.get('level', 0)
                lv_lbl = ['不足标准1', '标准1', '标准2', '标准3', '标准4', '标准5'][lv] if lv <= 5 else str(lv)
                # 改派标记：档位格子中显示改派数量
                _tf = d.get('transfer_orders', 0)
                if _tf > 0:
                    lv_lbl = f'{lv_lbl}(扣{_tf:.0f}单)'
                dlv = d.get('delivering', 0)

                write_data(ws4, r4, [d['name'], lv_lbl, ot,
                                       f'{dlv:.0f}单' if dlv > 0 else '无',
                                       f'{d.get("effective_orders", d["orders"]):.0f}单',
                                       fmt_time(d["online"]),
                                       '、'.join(gp) if gp else '已达标',
                                       d.get('remaining_shifts', '无'),
                                       can_achieve_yn(d.get('can_achieve', ''))])
                # 行级着色
                _c4=str(ws4.cell(r4,9).value or '')
                _g4=str(ws4.cell(r4,7).value or '')
                if _c4=='是':
                    # 能达成=绿底
                    for _cc in range(1,10):ws4.cell(r4,_cc).fill=PatternFill(start_color='F0FFF4',end_color='F0FFF4',fill_type='solid')
                    ws4.cell(r4,9).font=Font(size=10,color='38A169',name='PingFang SC')
                elif '1单'in _g4:
                    # 差1单=琥珀底，需关注
                    for _cc in range(1,10):ws4.cell(r4,_cc).fill=PatternFill(start_color='FFFAF0',end_color='FFFAF0',fill_type='solid')
                    ws4.cell(r4,7).font=Font(size=10,color='D69E2E',name='PingFang SC')
                    ws4.cell(r4,9).font=Font(size=10,color='718096',name='PingFang SC')
                else:
                    # 默认
                    ws4.cell(r4,7).font=Font(size=10,color='E53E3E',name='PingFang SC')
                # 改派着色（放最后，覆盖档位格子）
                if _tf > 0:
                    ws4.cell(r4,2).fill=PatternFill(start_color='FFF3CD',end_color='FFF3CD',fill_type='solid')
                    ws4.cell(r4,2).font=Font(size=10,bold=True,color='C05621',name='PingFang SC')
                    ws4.cell(r4,2).alignment=Alignment(horizontal='center',vertical='center')
                r4 += 1

        r4 += 1
        lv3_plus = rf.get('lv3_plus', 0)
        below_count = rf.get('below_count', 0)
        potential = lv3_plus + len(lv3_candidates)

        # 汇总信息分两行显示
        summary_line1 = f'  📊 数据汇总'
        summary_line2 = f'  标准3以上: {lv3_plus}人    |    升档机会: {total_up}人    |    潜力可达: {potential}人    |    未达成一档: {below_count}人'

        r4 += 1
        ws4.merge_cells(f'A{r4}:I{r4}')
        ws4.cell(r4, 1).value = summary_line1
        ws4.cell(r4, 1).font = Font(size=11, bold=True, color=C_PRIMARY, name='PingFang SC')
        ws4.cell(r4, 1).fill = FILL_SECTION
        ws4.cell(r4, 1).alignment = ALIGN_LEFT
        for cx in range(2, 10):
            ws4.cell(r4, cx).fill = FILL_SECTION
        set_row(ws4, r4, 30)

        r4 += 1
        ws4.merge_cells(f'A{r4}:I{r4}')
        ws4.cell(r4, 1).value = summary_line2
        ws4.cell(r4, 1).font = Font(size=11, bold=False, color=C_TEXT, name='PingFang SC')
        ws4.cell(r4, 1).fill = FILL_BG
        ws4.cell(r4, 1).alignment = ALIGN_LEFT
        for cx in range(2, 10):
            ws4.cell(r4, cx).fill = FILL_BG
        set_row(ws4, r4, 28)

        # ── 尖峰场景汇总 ──
        if rl:
            _sc_level_counts = {i: 0 for i in range(6)}
            for _sc_rv in rl.values():
                if _sc_rv['orders'] >= BASELINE_ORDERS and _sc_rv['online'] >= BASELINE_HOURS:
                    _sc_level_counts[_sc_rv['level']] += 1
            _sc_total_active = sum(1 for _sc_rv in rl.values() if _sc_rv['effective_orders'] > 0)
            _sc_b2b5 = _sc_level_counts.get(2,0) + _sc_level_counts.get(3,0) + _sc_level_counts.get(4,0) + _sc_level_counts.get(5,0)
            _sc_b2b5_pct = _sc_b2b5 / _target_capacity * 100 if _target_capacity else 0
            _sc_is_bad_weather = _weather_level in ('轻微恶天','一般恶天','严重恶天')
            _sc_is_severe = _weather_level == '严重恶天'
            _sc_is_peak = _is_weekend or _sc_is_bad_weather
            _sc_jf_score = 2 if _sc_is_severe else 1
            if _sc_is_peak:
                _sc_jf_label = f'尖峰场景（{"周末 " if _is_weekend else ""}{"严重恶天" if _sc_is_severe else "恶天" if _sc_is_bad_weather else ""}）'
            else:
                _sc_jf_label = '非尖峰场景'
            if not _sc_is_peak:
                _sc_jf_result = '非尖峰场景，不考核'
            _sc_need_70 = int(_target_capacity * 0.70) + 1
            _sc_gap_70 = max(0, _sc_need_70 - _sc_b2b5)
            if _sc_b2b5_pct < 30:
                _sc_jf_result = f'⚠️ 标2-标5占比＜30%，当天扣{_sc_jf_score}分'
            elif _sc_b2b5_pct > 70:
                _sc_jf_result = f'✅ 标2-标5占比＞70%，当天加{_sc_jf_score}分'
            else:
                _sc_jf_result = f'➖ 不加不扣 | 还需{_sc_gap_70}人上标2以上→今日加{_sc_jf_score}分'
            r4 += 1
            ws4.merge_cells(f'A{r4}:I{r4}')
            ws4.cell(r4, 1).value = f'🌤 天气: {_weather_info} | {_weather_level} | {_day_name} | {_sc_jf_label} | 标2-标5: {_sc_b2b5}/{int(_target_capacity)}={_sc_b2b5_pct:.1f}% | {_sc_jf_result}'
            ws4.cell(r4, 1).font = Font(size=10, color=C_MUTED, name='PingFang SC')
            ws4.cell(r4, 1).fill = FILL_BG
            ws4.cell(r4, 1).alignment = ALIGN_LEFT
            for cx in range(2, 10):
                ws4.cell(r4, cx).fill = FILL_BG
            set_row(ws4, r4, 22)
        ws4.cell(r4, 1).font = Font(size=10, color=C_MUTED, name='PingFang SC')
        ws4.cell(r4, 1).fill = FILL_BG
        ws4.cell(r4, 1).alignment = ALIGN_LEFT
        for cx in range(2, 10):
            ws4.cell(r4, cx).fill = FILL_BG
        set_row(ws4, r4, 22)

        # ── 未达成一档（参考原版配色：蓝底/红底）──
        bd = rf.get('below_details', {})
        if bd:
            only_time = []; only_order = []; both = []
            for _d in bd.values():
                _go = _d.get('gap_orders', 0); _gh = _d.get('gap_hours', 0)
                if _go > 0 and _gh > 0: both.append(_d)
                elif _go > 0: only_order.append(_d)
                else: only_time.append(_d)
            only_time.sort(key=lambda x: x.get('gap_hours', 0))
            only_order.sort(key=lambda x: (x.get('gap_orders', 0), -x.get('effective_orders', 0)))
            both.sort(key=lambda x: (x.get('gap_orders', 0) + x.get('gap_hours', 0), -x.get('effective_orders', 0)))

            for _title, _list, _fa, _fb, _fc, _gc in [
                ('🕐 只缺时长  （单量≥18单，在线<6h）共{}人', only_time, 'F7FAFC', 'EDF2F7', '2D3748', '2D3748'),
                ('📦 只缺单量  （在线≥6h，单量<18单）共{}人', only_order, 'EBF8FF', 'F7FAFC', '3182CE', '3182CE'),
                ('🔴 都缺  （单量<18单 且 在线<6h）共{}人', both, 'FFF5F5', 'F7FAFC', 'E53E3E', 'E53E3E'),
            ]:
                if not _list: continue
                r4 += 1
                write_section(ws4, r4, _title.format(len(_list)))
                r4 += 1
                write_header(ws4, r4, ['骑手', '实完单', '有效时长', '差距', '后续班次', '能否达成'])
                r4 += 1
                for _i, _d in enumerate(_list):
                    _go = _d.get('gap_orders', 0); _gh = _d.get('gap_hours', 0)
                    _gp = []
                    if _go > 0: _gp.append(f'差{_go:.0f}单')
                    if _gh > 0: _gp.append(f'差{fmt_time(_gh)}')
                    _tf = _d.get('transfer_orders', 0)
                    _name_display = _d.get('name', '')
                    if _tf > 0:
                        _name_display = f'{_name_display}(扣{_tf:.0f}单)'
                    _fill = PatternFill(start_color=_fa, end_color=_fa, fill_type='solid') if _i % 2 == 0 else PatternFill(start_color=_fb, end_color=_fb, fill_type='solid')
                    _fnt = Font(size=10, color=_fc, name='PingFang SC')
                    write_data(ws4, r4, [
                        _name_display,
                        f'{_d.get("effective_orders", _d.get("orders", 0)):.0f}单',
                        fmt_time(_d.get('online', 0)),
                        '、'.join(_gp) if _gp else '',
                        _d.get('remaining_shifts', '无'),
                        can_achieve_yn(_d.get('can_achieve', '')),
                    ], fill=_fill, font=_fnt)
                    # 改派着色：骑手格子变橙底
                    if _tf > 0:
                        ws4.cell(r4,1).fill=PatternFill(start_color='FFF3CD',end_color='FFF3CD',fill_type='solid')
                        ws4.cell(r4,1).font=Font(size=10,bold=True,color='C05621',name='PingFang SC')
                    # 差距列着色
                    _gv4 = str(ws4.cell(r4, 4).value or '')
                    if '1单' in _gv4: ws4.cell(r4, 4).font = Font(size=10, color='D69E2E', name='PingFang SC')
                    elif _gv4 and _gv4 not in ['', 'None']: ws4.cell(r4, 4).font = Font(size=10, color='E53E3E', name='PingFang SC')
                    # 能否达成着色
                    _cv4 = str(ws4.cell(r4, 6).value or '')
                    if _cv4 == '是': ws4.cell(r4, 6).font = Font(size=10, color='38A169', name='PingFang SC')
                    elif _cv4 == '否': ws4.cell(r4, 6).font = Font(size=10, color=_gc, name='PingFang SC')
                    r4 += 1

        auto_fit_cols(ws4, 9)

    # ═══════════════════════════════════════
    # 💾 保存
    # ═══════════════════════════════════════

    # ═══ Sheet 6: 0-2档不够明细 ═══
    # 目标档位定义 (level → next_level 的目标单量/时长)
    NEXT_LEVEL = {
        0: {'orders': 18, 'hours': 6,  'label': '标准1'},
        1: {'orders': 28, 'hours': 6,  'label': '标准2'},
        2: {'orders': 35, 'hours': 6,  'label': '标准3'},
    }
    ws6 = wb.create_sheet('0-2档不够明细')
    ws6.merge_cells('A1:J1')
    ws6['A1'].value = f'0-2档不够明细  {sd}  {_data_time}'
    ws6['A1'].font = FONT_TITLE
    ws6['A1'].fill = FILL_HEADER
    ws6['A1'].alignment = ALIGN_LEFT
    for c in range(2, 11): ws6.cell(1, c).fill = FILL_HEADER
    set_row(ws6, 1, 40)
    # A列宽度由行尾部 auto_fit 自适应
    if rl:
        riders_lv012 = [rv for rv in rl.values() if rv['level'] in (0, 1)]
        ws6.cell(2, 1).value = f'当前0-1档不够骑手共 {len(riders_lv012)} 人'
        ws6.cell(2, 1).font = FONT_MUTED
        ws6.cell(2, 1).alignment = ALIGN_LEFT
        set_row(ws6, 2, 22)

        # 第3行：场景信息
        _s1_scene_line6 = f'🌤 {_weather_info}/{_weather_level} | {_day_name} | 尖峰场景（{"周末 " if _is_weekend else ""}{"严重恶天" if _s1_is_severe else "恶天" if _s1_is_bad else ""}）'
        if _s1_b2b5_pct < 30:
            _s1_score_line6 = f'当前{_s1_b2b5_pct:.1f}% | ⚠️ 低于30% → 扣{_s1_jf_score}分'
        elif _s1_b2b5_pct > 70:
            _s1_score_line6 = f'当前{_s1_b2b5_pct:.1f}% | ✅ 超过70% → 加{_s1_jf_score}分'
        else:
            _s1_score_line6 = f'当前{_s1_b2b5_pct:.1f}% | 还差{_s1_gap_70}人达标2'
        _s1_transfer_line6 = f'📋 改派扣减合计: {_total_transfer}单' if _total_transfer > 0 else '📋 无扣减'
        for _s6_ri, _s6_txt in enumerate([_s1_scene_line6, _s1_score_line6, _s1_transfer_line6], start=3):
            ws6.merge_cells(f'A{_s6_ri}:J{_s6_ri}')
            ws6.cell(_s6_ri, 1).value = _s6_txt
            ws6.cell(_s6_ri, 1).font = Font(size=11, bold=True, color=C_TEXT, name='PingFang SC')
            ws6.cell(_s6_ri, 1).alignment = ALIGN_LEFT
            ws6.cell(_s6_ri, 1).fill = FILL_SECTION
            for _scx in range(2, 11): ws6.cell(_s6_ri, _scx).fill = FILL_SECTION
            set_row(ws6, _s6_ri, 22)

        # A7 分类统计
        _lv0 = [rv for rv in riders_lv012 if rv['level'] == 0]
        _lv1 = [rv for rv in riders_lv012 if rv['level'] == 1]
        ws6.merge_cells(f'A7:J7')
        ws6.cell(7, 1).value = f'不够 {len(riders_lv012)}人 | 未达标 {len(_lv0)}人 | 标准1 {len(_lv1)}人 | 已达标2档已过滤'
        ws6.cell(7, 1).font = FONT_BOLD
        ws6.cell(7, 1).alignment = ALIGN_LEFT
        for _cx in range(2, 11): ws6.cell(7, _cx).fill = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
        set_row(ws6, 7, 22)

        r6 = 8
        write_header(ws6, r6, ['骑手', '当前→晋升', '有效完单', '有效时长', '配送中单量', '是否在线', '差单', '差时', '最后班次', '能否达成'])
        r6 += 1
        for rv in sorted(riders_lv012, key=lambda x: (x['level'], -x['effective_orders'], -x['online'], x['name'])):
            eff = rv['effective_orders']
            on = rv['online']
            lv = rv['level']
            tgt = NEXT_LEVEL[1]  # 统一算到标准2（28单/6h）
            tgt_orders = tgt['orders']
            tgt_hours = tgt['hours']
            go = max(0, tgt_orders - eff)
            gh = max(0, round(tgt_hours - on, 1))
            ot = '🟢 在线' if is_online(rv.get('work_status', '')) else '🔴 离线'
            _sched6 = rv.get('shift_schedule', '')
            rs = fc_mod.get_remaining_periods(_sched6, current_hour=data_hour)
            _sched6_parts = [p.strip() for p in _sched6.replace('，',',').replace('、',',').split(',') if p.strip()]
            _last6 = _sched6_parts[-1] if _sched6_parts else '—'
            ca = fc_mod.estimate_achieve(eff, on, tgt_orders, rs, rv.get('delivering', 0), current_hour=data_hour, target_hours=tgt_hours)
            _tf6 = rv.get('transfer_orders', 0)
            _name6 = rv['name']
            if _tf6 > 0:
                _name6 = f'{_name6}(扣{_tf6:.0f}单)'
            # 当前档位标签
            _lv_labels = {0: '未达标', 1: '标准1', 2: '标准2'}
            _cur_label = _lv_labels.get(lv, f'标准{lv}')
            write_data(ws6, r6, [
                _name6,
                f'{_cur_label}→标准2',
                f'{eff:.0f}单',
                fmt_time(on),
                f'{rv.get("delivering", 0):.0f}单',
                ot,
                '-' if go <= 0 else f'{go:.0f}',
                '-' if gh <= 0 else f'-{fmt_time(gh)}',
                _last6,
                can_achieve_yn(ca),
            ])
            # 骑手列居中
            ws6.cell(r6,1).alignment=Alignment(horizontal='center',vertical='center')
            # 改派着色：骑手单元格橙底
            if _tf6 > 0:
                ws6.cell(r6,1).fill=PatternFill(start_color='FFF3CD',end_color='FFF3CD',fill_type='solid')
                ws6.cell(r6,1).font=Font(size=10,bold=True,color='C05621',name='PingFang SC')
            # 差单标红、差时标红
            if str(ws6.cell(r6,7).value or '') not in ['', '-', 'None']: ws6.cell(r6,7).font=FONT_FAIL
            if str(ws6.cell(r6,8).value or '') not in ['', '-', 'None']: ws6.cell(r6,8).font=FONT_FAIL
            # 能否达成：是=绿 否=灰（参考原版）
            _ca2 = str(ws6.cell(r6,10).value or '')
            if _ca2=='是':
                for _cc2 in range(1,11):ws6.cell(r6,_cc2).fill=PatternFill(start_color='F0FFF4',end_color='F0FFF4',fill_type='solid')
                ws6.cell(r6,10).font=Font(size=10,color='38A169',name='PingFang SC')
            elif _ca2=='否':
                ws6.cell(r6,10).font=Font(size=10,color='718096',name='PingFang SC')
            r6 += 1
        auto_fit_cols(ws6, 10)
        # 骑手列（A列）单独自适应（排除标题/汇总行，只算数据行）
        _max_a = 10
        for _r6 in range(9, ws6.max_row + 1):
            _v6 = ws6.cell(_r6, 1).value
            if _v6:
                _w6 = sum(2 if ord(ch) > 127 else 1 for ch in str(_v6))
                if _w6 > _max_a: _max_a = _w6
        ws6.column_dimensions['A'].width = min(_max_a + 2, 40)



    # ═══════════════════════════════════════
    # 📊 档位明细（9列：姓名 | 档位 | 有效时长 | 全天完单 | 配送中 | 档位标准 | 差单 | 晋升档位 | 最后班次）
    # 配送中不算当天完成量
    # ═══════════════════════════════════════
    
    ws7 = wb.create_sheet('档位明细')

    NEXT_TGT = [
        (0, 18, 6),
        (1, 28, 6),
        (2, 35, 6),
        (3, 42, 7),
        (4, 53, 8),
    ]
    LVL_NAMES = ['不足标准1', '标准1', '标准2', '标准3', '标准4', '标准5']
    NEXT_NAMES = ['1档', '2档', '3档', '4档', '5档', '已到顶']
    FILL_GREEN = PatternFill(start_color='D4EDDA',end_color='D4EDDA',fill_type='solid')
    FILL_ORANGE = PatternFill(start_color='FFE0B2',end_color='FFE0B2',fill_type='solid')
    FILL_GAP12 = PatternFill(start_color='FFF9C4',end_color='FFF9C4',fill_type='solid')

    _nc = 11  # 姓名 状态 档位 有效时长 全天完单 配送中 档位标准 差单 晋升档位 最后班次 班次差单
    ws7.merge_cells(f'A1:{get_column_letter(_nc)}1')
    ws7['A1'].value = f'档位明细  {sd}  {_data_time}'
    ws7['A1'].font = FONT_TITLE
    ws7['A1'].fill = FILL_HEADER
    ws7['A1'].alignment = ALIGN_LEFT
    for c in range(2, _nc + 1): ws7.cell(1, c).fill = FILL_HEADER
    set_row(ws7, 1, 40)

    if rl:
        _s7r = 2
        for _s7_txt in [
            f'  🌤 {_weather_info}/{_weather_level} | {_day_name} | {_sc_jf_label}',
            f'  当前{_sc_b2b5_pct:.1f}% | {_sc_jf_result}',
            f'  📋 改派扣减合计: {_total_transfer}单' if _total_transfer > 0 else '  📋 无扣减',
        ]:
            ws7.merge_cells(f'A{_s7r}:{get_column_letter(_nc)}{_s7r}')
            ws7.cell(_s7r, 1).value = _s7_txt
            ws7.cell(_s7r, 1).font = Font(size=10, color=C_TEXT, name='PingFang SC')
            ws7.cell(_s7r, 1).alignment = ALIGN_LEFT
            for _scx in range(2, _nc + 1): ws7.cell(_s7r,_scx).fill = FILL_SECTION
            set_row(ws7, _s7r, 22)
            _s7r += 1

        _dist_parts = [f'{LVL_NAMES[_lv]} {_sc_level_counts.get(_lv,0)}人' for _lv in [5,4,3,2,1,0]]
        ws7.merge_cells(f'A{_s7r}:{get_column_letter(_nc)}{_s7r}')
        ws7.cell(_s7r,1).value = f'  {' | '.join(_dist_parts)}'
        ws7.cell(_s7r,1).font = Font(size=10, bold=True, color=C_PRIMARY, name='PingFang SC')
        ws7.cell(_s7r,1).alignment = ALIGN_LEFT
        for _scx in range(2, _nc + 1): ws7.cell(_s7r,_scx).fill = FILL_SECTION
        set_row(ws7, _s7r, 22)
        _s7r += 1

        write_header(ws7, _s7r, ['姓名', '状态', '档位', '有效时长', '全天完单', '配送中', '档位标准', '差单', '晋升档位', '最后班次', '班次差单'])
        for _cc in range(1, _nc + 1):
            ws7.cell(_s7r,_cc).border = Border(
                left=Side(style='thin', color='1A365D'),
                right=Side(style='thin', color='1A365D'),
                top=Side(style='thin', color='1A365D'),
                bottom=Side(style='medium', color='1A365D'))
        _s7r += 1

        # ── 构建骑手班次差单查询（与时段未达标同源数据）──
        _rider_period_gap = {}  # (name, period) -> 缺口单量
        if isinstance(ra, tuple) and ra[0]:
            for _pd in ra[0]:
                _rn = _pd.get('name', '')
                _pp = _pd.get('period', '')
                _gp = _pd.get('gap_order', 0)  # gap_order = 实完单 - 要求单（负=缺单）
                _rider_period_gap[(_rn, _pp)] = _gp

        # ── 确定当前进行中时段 ──
        _current_period = None
        for _pn in period_order:
            _ps, _pe = PERIOD_HOURS.get(_pn, (0, 24))
            if _ps <= data_hour < _pe:
                _current_period = _pn
                break

        _baseline_riders = [rv for rv in rl.values() if rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS]
        _below_riders = [rv for rv in rl.values() if not (rv['orders'] >= BASELINE_ORDERS and rv['online'] >= BASELINE_HOURS)]
        _lv_groups = {0: _below_riders}
        for rv in _baseline_riders:
            _lv_groups.setdefault(rv['level'], []).append(rv)

        for _lv in [4, 3, 2, 1, 0]:  # 只显示0~4档，已达顶的标5不展示
            _riders = _lv_groups.get(_lv, [])
            if not _riders: continue

            for _ri, rv in enumerate(sorted(_riders, key=lambda x: (-x.get('effective_orders', x['orders']), -x['online']))):
                _nm = rv.get('name', '')
                _eff = rv.get('effective_orders', rv['orders'])
                _on = rv['online']
                _dlv = rv.get('delivering', 0)
                _tf7 = rv.get('transfer_orders', 0)
                _sched_str = rv.get('shift_schedule', '')
                _rs = fc_mod.get_remaining_periods(_sched_str, current_hour=data_hour)
                _sched_parts = [p.strip() for p in _sched_str.replace('，',',').replace('、',',').split(',') if p.strip()]
                _last_shift = _sched_parts[-1] if _sched_parts else '—'
                _nm7 = _nm if _tf7 == 0 else f'{_nm}(扣{_tf7:.0f}单)'

                # 档位标准 + 差单
                if _lv < 5:
                    _tgt_o = NEXT_TGT[_lv][1]
                    _tgt_h = NEXT_TGT[_lv][2]
                    _std_str = f'{_lv+1}档{_tgt_o}'
                    _go = max(0, _tgt_o - _eff)
                    _go_str = f'{_go:.0f}' if _go > 0 else '-'
                else:
                    _std_str = '已达顶'
                    _go = 0
                    _go_str = '-'

                _dlv_str = f'{_dlv:.0f}单' if _dlv > 0 else '-'

                # 当前差单：仅当骑手排班包含当前时段才显示，否则—
                if _current_period and _current_period in _sched_parts:
                    _curr_gap = _rider_period_gap.get((_nm, _current_period), None)
                    if _curr_gap is not None and _curr_gap < 0:
                        _curr_gap_str = f'{-_curr_gap:.0f}单'
                    else:
                        _curr_gap_str = '✓'
                else:
                    _curr_gap_str = '—'

                write_data(ws7, _s7r, [
                    _nm7, format_status(rv.get('work_status', '')), LVL_NAMES[_lv], fmt_time(_on),
                    f'{_eff:.0f}单', _dlv_str,
                    _std_str, _go_str, NEXT_NAMES[_lv],
                    _last_shift, _curr_gap_str,
                ])

                # 底色：差1-2浅黄 → 改派橙 → 白灰斑马
                _is_gap12 = 0 < _go <= 2
                if _is_gap12:
                    _row_fill = FILL_GAP12
                elif _tf7 > 0:
                    _row_fill = FILL_ORANGE
                else:
                    _row_fill = FILL_ALT if _ri % 2 == 1 else FILL_BG
                for _cc in range(1, _nc + 1): ws7.cell(_s7r,_cc).fill = _row_fill
                ws7.cell(_s7r,1).alignment = Alignment(horizontal='center', vertical='center')

                # 改派标橙
                if _tf7 > 0:
                    ws7.cell(_s7r,1).font=Font(size=10,bold=True,color='C05621',name='PingFang SC')

                # 差单着色（第8列）
                if _go > 0:
                    if _is_gap12:
                        ws7.cell(_s7r,8).font=Font(size=10,bold=True,color='D84315',name='PingFang SC')
                    else:
                        ws7.cell(_s7r,8).font=Font(size=10,color='E53E3E',name='PingFang SC')
                else:
                    ws7.cell(_s7r,8).font=Font(size=10,color='38A169',name='PingFang SC')

                # 当前差单：不用标红，不排班就灰色，其他正常
                if _curr_gap_str == '—':
                    ws7.cell(_s7r,11).font=Font(size=10,color='CBD5E0',name='PingFang SC')
                else:
                    ws7.cell(_s7r,11).font=Font(size=10,color='718096',name='PingFang SC')

                _s7r += 1

        _lv3_plus7 = sum(1 for rv in _baseline_riders if rv.get('level',0) >= 3)
        _below7 = len(_below_riders)
        ws7.merge_cells(f'A{_s7r}:{get_column_letter(_nc)}{_s7r}')
        ws7.cell(_s7r,1).value = f'  标准3以上: {_lv3_plus7}人  |  低于基线: {_below7}人'
        ws7.cell(_s7r,1).font = Font(size=11, bold=True, color=C_PRIMARY, name='PingFang SC')
        ws7.cell(_s7r,1).fill = FILL_SECTION; ws7.cell(_s7r,1).alignment = ALIGN_LEFT
        for _cxx in range(2, _nc + 1): ws7.cell(_s7r,_cxx).fill = FILL_SECTION
        set_row(ws7, _s7r, 28)

        _col_widths = {'A':14,'B':10,'C':12,'D':11,'E':10,'F':11,'G':12,'H':8,'I':8,'J':10,'K':10}
        for _r7a in range(2, ws7.max_row):
            for _cc in range(1, _nc + 1):
                _c7v = ws7.cell(_r7a, _cc).value
                if _c7v:
                    _cl = get_column_letter(_cc)
                    _w7b = sum(2 if ord(ch) > 127 else 1 for ch in str(_c7v))
                    if _w7b > _col_widths[_cl]: _col_widths[_cl] = min(_w7b + 2, 35)
        _col_widths['A'] = min(_col_widths['A'], 22)
        for _cl, _w7 in _col_widths.items():
            ws7.column_dimensions[_cl].width = _w7
    fp_out = Path.home() / 'Desktop' / f'配送日报-{sd[5:]}.xlsx'
    wb.save(str(fp_out))
    print(f'✅ {fp_out}')


if __name__ == '__main__':
    gen(sys.argv[1] if len(sys.argv) > 1 else None)

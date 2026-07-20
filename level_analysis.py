#!/usr/bin/env python3
"""
档位达成分析 — 读取出勤 Excel，计算每个骑手的档位和分布
"""
import os, re, json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system('pip3 install openpyxl --break-system-packages >/dev/null 2>&1')
    import openpyxl

LEVELS = [
    {'level': 1, 'orders': 18, 'hours': 6,  'label': '标准1:6h/18单'},
    {'level': 2, 'orders': 28, 'hours': 6,  'label': '标准2:6h/28单'},
    {'level': 3, 'orders': 35, 'hours': 6,  'label': '标准3:6h/35单'},
    {'level': 4, 'orders': 42, 'hours': 7,  'label': '标准4:7h/42单'},
    {'level': 5, 'orders': 53, 'hours': 8,  'label': '标准5:8h/53单'},
]

def _is_assessable(配送标品, 商家订单流水号, 是否逆向单, 欺诈是否成立):
    """判断订单是否计入考核（是否考核=1）
    
    与监控表 运单明细!C列（是否考核）公式逻辑一致：
    =IF(M="","",IF(
        FIND("帮买",配送标品)+FIND("帮送",配送标品)+FIND("淘宝逆向单",配送标品)+FIND("淘鲜达",配送标品)
        +FIND("高校",配送标品)+FIND("波次达",配送标品)
        +IF(是否逆向单="是",1,0)+IF(欺诈是否成立="是",1,0)
        +FIND("抖音",商家订单流水号)+FIND("美团",商家订单流水号)
        =0, 1, 0))
    返回 True = 考核, False = 不考核
    """
    # 配送标品排除项
    标品 = str(配送标品 or '')
    for kw in ['帮买', '帮送', '淘宝逆向单', '淘鲜达', '高校', '波次达']:
        if kw in 标品:
            return False
    if str(是否逆向单 or '').strip() == '是':
        return False
    if str(欺诈是否成立 or '').strip() == '是':
        return False
    # 检查商家订单流水号是否含 抖音/美团
    流水号 = str(商家订单流水号 or '')
    if '抖音' in 流水号 or '美团' in 流水号:
        return False
    return True


def _is_same_day(骑手接单时间, 运单完成时间):
    """判断是否为当日订单（当日判定=1）
    
    与监控表 运单明细!D列（当日判定）公式逻辑一致：
    =IF(M="","",IFERROR(IF(MID(AP,9,2)-MID(AK,9,2)=0,1,0),1))
    比较运单完成时间和骑手接单时间的日期是否相同
    """
    try:
        if not 骑手接单时间 or not 运单完成时间:
            return False
        t1 = str(骑手接单时间).strip()
        t2 = str(运单完成时间).strip()
        # 提取 YYYY-MM-DD 部分的 DD（日期）
        d1 = t1[8:10] if len(t1) >= 10 else t1[:2]
        d2 = t2[8:10] if len(t2) >= 10 else t2[:2]
        return d1 == d2
    except Exception:
        return True  # 解析失败默认当日


def read_transfer_data():
    """读取最新运单数据
    
    返回 (combined, transfer_only) 两个字典：
    - combined: {骑手名称: 总扣减量}（改派+无效，与监控表口径一致，用于档位计算）
    - transfer_only: {骑手名称: 改派单量}（仅改派，用于日报显示）
    
    扣减逻辑与监控表 实时考勤!A列 口径一致：
    - 无效单：不考核 OR 非当日 的订单
    - 改派单：是否发生改派=是（且非配送失败）
    - 一单不会同时计入改派和无效（互斥）
    """
    files = sorted(Path.home().glob('Downloads/*运单*.xlsx'),
                   key=lambda f: os.path.getmtime(f), reverse=True)
    files = [f for f in files if not f.name.startswith('.') and not f.name.startswith('~') and f.stat().st_size > 10000]
    if not files:
        return {}, {}
    try:
        wb = openpyxl.load_workbook(str(files[0]), data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        def find_col(targets):
            for t in targets:
                for i, h in enumerate(headers):
                    if h and t in str(h):
                        return i + 1
            return None

        name_col = find_col(['骑手名称'])
        if not name_col:
            wb.close()
            return {}, {}

        transfer_col = find_col(['是否发生改派'])
        delivery_col = find_col(['配送标品'])
        order_id_col = find_col(['商家订单流水号'])
        reverse_col = find_col(['是否逆向单'])
        fraud_col = find_col(['欺诈是否成立'])
        accept_col = find_col(['骑手接单时间'])
        finish_col = find_col(['运单完成时间'])
        status_col = find_col(['运单状态'])

        combined = {}     # 改派+无效（用于内部计算）
        transfer_only = {}  # 仅改派（用于日报显示）
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            if not name:
                continue

            # ── 无效单量判定（与监控表公式一致） ──
            is_valid = True
            if delivery_col and order_id_col:
                d = ws.cell(r, delivery_col).value
                oid = ws.cell(r, order_id_col).value
                rev = ws.cell(r, reverse_col).value if reverse_col else None
                fra = ws.cell(r, fraud_col).value if fraud_col else None
                if not _is_assessable(d, oid, rev, fra):
                    is_valid = False

            if is_valid and accept_col and finish_col:
                accept_t = ws.cell(r, accept_col).value
                finish_t = ws.cell(r, finish_col).value
                if not _is_same_day(accept_t, finish_t):
                    is_valid = False

            if not is_valid:
                # 无效单：不考核 OR 非当日（仅计入combined，不计入transfer_only）
                combined[name] = combined.get(name, 0) + 1
                continue

            # ── 改派单量判定 ──
            if transfer_col and status_col:
                val = ws.cell(r, transfer_col).value
                status = str(ws.cell(r, status_col).value or '').strip()
                # 与监控表改派公式一致：非配送失败 + 是否发生改派=是
                if val is not None and str(val).strip() == '是' and status != '配送失败':
                    combined[name] = combined.get(name, 0) + 1
                    transfer_only[name] = transfer_only.get(name, 0) + 1

        wb.close()
        return combined, transfer_only
    except Exception as e:
        print(f'⚠️ 读取运单数据失败: {e}')
        return {}, {}


def fmt_time(h):
    if h is None or h == 0: return '0分钟'
    total_m = round(h * 60)
    hrs, mins = divmod(total_m, 60)
    if hrs == 0: return f'{mins}分钟'
    if mins == 0: return f'{hrs}小时'
    return f'{hrs}小时{mins}分钟'

def parse_time(t):
    if not t or t == '-' or t is None: return 0.0
    # 纯数字→秒（新格式：全天有效在线时长/时段有效在线时长以秒存储）
    if isinstance(t, (int, float)):
        return t / 3600.0
    t = str(t).strip()
    h = 0.0; m = 0.0
    g = re.search(r'(\d+)小时', t)
    if g: h = float(g.group(1))
    g = re.search(r'(\d+)分钟', t)
    if g: m = float(g.group(1))
    if not h and not m:
        g = re.search(r'^(\d+):(\d+)$', t)
        if g: h, m = float(g.group(1)), float(g.group(2))
    return h + m / 60

def calc_level(orders, hours):
    """计算档位（取低原则）"""
    order_lv = 0
    hours_lv = 0
    for lv in LEVELS:
        if orders >= lv['orders']: order_lv = lv['level']
        if hours >= lv['hours']:   hours_lv = lv['level']
    return min(order_lv, hours_lv) if order_lv and hours_lv else 0

def analyze(filepath=None, transfer_data=None, transfer_only_data=None):
    if not filepath:
        files = sorted(Path.home().glob('Downloads/*出勤*.xlsx'),
                       key=lambda f: os.path.getmtime(f), reverse=True)
        files = [f for f in files if not f.name.startswith('.') and not f.name.startswith('~') and f.stat().st_size > 10000]
        if not files:
            print("❌ 未找到出勤文件")
            return
        filepath = files[0]

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def find_col(targets, blacklist=None):
        for t in targets:
            for i, h in enumerate(headers):
                if h and t in str(h) and not (blacklist and any(b in str(h) for b in blacklist)):
                    return i + 1
        return None

    col = {
        '站点': find_col(['站点名称']),
        '姓名': find_col(['姓名', '骑手姓名']),
        '时段': find_col(['时段'], blacklist=['排班']),
        '排班': find_col(['排班时段', '排班状态', '是否排班']),
        '完单': find_col(['全天完单量', '完单量'], blacklist=['目标']),
        '时长': find_col(['全天有效在线时长', '有效在线时长'], blacklist=['目标']),
        '配送中': find_col(['配送中单量']),
        '工作': find_col(['工作状态']),
    }

    if not col['姓名']:
        print(f"❌ 无法识别列结构: {headers[:20]}")
        return

    station = None
    results = {}
    seen = set()

    is_period_col = col.get('时段') is not None  # 有时段列=明细格式，无时段列=汇总格式
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col['姓名']).value
        if not name: continue
        if name in seen: continue

        if is_period_col:
            period = ws.cell(r, col['时段']).value
            # 时段明细格式：跳过非「全天」的行（各时段明细），只取全天汇总
            if not period or str(period).strip() != '全天': continue
        # 无时段列：每行都是骑手汇总数据，全部处理
        seen.add(name)

        s = ws.cell(r, col['站点']).value if col['站点'] else None
        if s: station = s

        orders_raw = ws.cell(r, col['完单']).value if col['完单'] else 0
        try: orders = float(orders_raw or 0)
        except: orders = 0

        # 有效完成单 = 全天完单量 - 总扣减量(改派+无效)
        # 与监控表 实时考勤!A列 = 全天完单量-SUMIFS(无效单量)-SUMIFS(改派单量) 口径一致
        if transfer_data is None:
            transfer_data = read_transfer_data()[0]
        transfer_orders = transfer_data.get(name, 0)
        effective_orders = max(0, orders - transfer_orders)

        online_raw = ws.cell(r, col['时长']).value if col['时长'] else 0
        online = parse_time(online_raw)

        schedule = str(ws.cell(r, col['排班']).value or '') if col['排班'] else '正常'
        # 排班时段字段不存在时，直接用排班状态值
        if schedule.strip() in ['', '休息', '未排班', '请假']: continue

        level = calc_level(effective_orders, online)

        # 离下一档差距（基于有效完成单）
        next_orders = 15 if level == 0 else (LEVELS[level]['orders'] if level < 5 else 999)
        next_hours = 3 if level == 0 else (LEVELS[level]['hours'] if level < 5 else 999)
        gap_order = max(0, next_orders - effective_orders)
        gap_hour = max(0, round(next_hours - online, 1))

        # 配送中单量 & 工作状态
        delivering_raw = ws.cell(r, col['配送中']).value if col['配送中'] else 0
        try: delivering = float(delivering_raw or 0)
        except: delivering = 0

        work_status = str(ws.cell(r, col['工作']).value or '') if col['工作'] else ''

        results[name] = {
            'name': name, 'orders': orders,  # 原始完单量（全天达成基线用）
            'online': round(online, 2),
            'level': level, 'gap_order': gap_order, 'gap_hour': gap_hour,
            'delivering': delivering, 'work_status': work_status,
            'shift_schedule': schedule,
            'effective_orders': effective_orders,  # 有效完成单（档位判断用）
            'transfer_orders': transfer_orders,
            'transfer_only_orders': (transfer_only_data or {}).get(name, 0),
        }

    wb.close()
    if not station: station = '未知站点'

    # ===== 输出 =====
    print(f"\n{'='*60}")
    print(f"🏆 档位达成统计")
    print(f"   站点: {station}")
    print(f"   文件: {Path(filepath).name}")
    print(f"{'='*60}")

    level_counts = {i: 0 for i in range(6)}
    for r in results.values(): level_counts[r['level']] += 1

    # ===== 基线过滤：6h/18单 =====
    BASELINE_HOURS = 6
    BASELINE_ORDERS = 18
    baseline_riders = {n: r for n, r in results.items() if r['orders'] >= BASELINE_ORDERS and r['online'] >= BASELINE_HOURS}
    below_riders = {n: r for n, r in results.items() if not (r['orders'] >= BASELINE_ORDERS and r['online'] >= BASELINE_HOURS)}

    total = len(results)
    total_baseline = len(baseline_riders)
    below_count = len(below_riders)
    rate_3plus = sum(1 for r in baseline_riders.values() if r['level'] >= 3)

    print(f"\n  🔍 基线过滤：≥{BASELINE_HOURS}h & ≥{BASELINE_ORDERS}单")
    print(f"  排班骑手: {total}人 | 达到基线: {total_baseline}人 | 低于基线: {below_count}人")
    print(f"  3档以上达标率: {rate_3plus}/{total_baseline} = {rate_3plus/total_baseline*100:.1f}%")

    level_counts = {i: 0 for i in range(6)}
    for r in baseline_riders.values(): level_counts[r['level']] += 1

    print(f"\n  {'档位':<8} {'标准':<16} {'人数':>4} {'占比':>7}")
    print(f"  {'─'*38}")
    lvl_label = {5: '标准5', 4: '标准4', 3: '标准3', 2: '标准2', 1: '标准1', 0: '未达标'}
    for lv in [5, 4, 3, 2, 1, 0]:
        c = level_counts.get(lv, 0)
        std = LEVELS[lv-1]['label'] if lv > 0 else '不足标准1'
        print(f"  {lvl_label[lv]:<8} {std:<16} {c:>4}  {c/total_baseline*100:>5.1f}%" if total_baseline else f"  {lvl_label[lv]:<8} {std:<16} {c:>4}   0.0%")

    # 接近升级
    print(f"\n  📌 接近升级骑手（差1单或0.5h内）:")
    found = False
    for r in sorted(baseline_riders.values(), key=lambda x: (-x['level'], x['name'])):
        if r['level'] >= 5: continue
        if (0 < r['gap_order'] <= 1 or 0 < r['gap_hour'] <= 0.5):
            found = True
            lv_name = '未达标' if r['level'] == 0 else f'标准{r["level"]}'
            parts = []
            if 0 < r['gap_order'] <= 1: parts.append(f"差{r['gap_order']:.0f}单")
            if 0 < r['gap_hour'] <= 0.5: parts.append(f"差{fmt_time(r['gap_hour'])}")
            print(f"    {r['name']:<8} {lv_name} → {' '.join(parts)}")
    if not found: print(f"    （无）")

    # 低于基线汇总
    if below_count > 0:
        print(f"\n  📌 低于基线（<{BASELINE_HOURS}h 或 <{BASELINE_ORDERS}单）: {below_count}人")
        for r in sorted(below_riders.values(), key=lambda x: (-x['orders'], -x['online'])):
            parts = []
            if r.get('transfer_only_orders', 0) > 0:
                parts.append(f'扣减{r["transfer_only_orders"]:.0f}')
            tag = '  ' + ' '.join(parts) if parts else ''
            level_text = '未达标' if r['level'] == 0 else f'标准{r["level"]}'
            print(f"    {r['name']:<8} {r['orders']:>4.0f}单 {r['online']:>4.1f}h {level_text}{tag}")

    return results

if __name__ == '__main__':
    import sys
    analyze(sys.argv[1] if len(sys.argv) > 1 else None)

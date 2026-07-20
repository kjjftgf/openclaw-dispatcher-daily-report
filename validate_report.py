#!/usr/bin/env python3
"""
日报格式验证脚本
检查 ~/Desktop/配送日报_{today}.xlsx 是否符合固化标准
"""
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌ 缺少 openpyxl，请先安装")
    sys.exit(1)

today = datetime.now().strftime("%Y-%m-%d")
today_short = today[5:]  # 07-20
# 新格式 配送日报-07-20，再试其他格式
fp_exact = Path.home() / f"Desktop/配送日报-{today_short}.xlsx"
if fp_exact.exists():
    fp = fp_exact
else:
    today_short = today[5:]  # 07-20
    fp_exact2 = Path.home() / f"Desktop/配送日报_{today_short}.xlsx"
    if fp_exact2.exists():
        fp = fp_exact2
    else:
        fp_exact_old = Path.home() / f"Desktop/配送日报_{today}.xlsx"
        if fp_exact_old.exists():
            fp = fp_exact_old
        else:
            matches = sorted(Path.home().glob(f"Desktop/配送日报_{today}_*.xlsx"),
                             key=lambda f: f.stat().st_mtime, reverse=True)
            if not matches:
                print(f"❌ 未找到配送日报文件: Desktop/配送日报-{today_short}.xlsx")
                print("请先生成日报：cd ~/.openclaw/workspace/data-analyst && python3 generate_report.py")
                sys.exit(1)
            fp = matches[0]
print(f"📁 验证文件: {fp.name}")

wb = openpyxl.load_workbook(str(fp))
errors = []

# ── 检查 Sheet 名称 ──
expected_sheets = ["全天达成", "时段未达标", "档位达成", "重点跟进", "0-2档不够明细", "档位明细"]
actual_sheets = wb.sheetnames
for s in expected_sheets:
    if s not in actual_sheets:
        errors.append(f"缺少 Sheet: {s}")
for s in actual_sheets:
    if s not in expected_sheets:
        errors.append(f"多余 Sheet: {s}")

# ── 检查各 Sheet ──
for sn in expected_sheets:
    if sn not in actual_sheets:
        continue
    ws = wb[sn]
    r1 = ws.cell(1, 1).value or ""

    if sn == "全天达成":
        if "全天达成" not in str(r1):
            errors.append(f"[全天达成] 标题不对: {r1}")
        # 检查表头是否存在（信息行在第5行，第2-4行为场景/比例/改派行）
        row2 = ws.cell(2, 1).value or ""
        row3 = ws.cell(3, 1).value or ""
        row4 = ws.cell(4, 1).value or ""
        row5 = ws.cell(5, 1).value or ""
        if "站点人数" in str(row5) and ("🌤" in str(row2) or "尖峰" in str(row2)):
            pass  # 有场景行（row2）+比例行（row3）+改派行（row4），信息行在row5
        elif "站点人数" in str(row4) and "🌤" in str(row2) and ("当前" in str(row3) or "%" in str(row3)):
            pass  # 旧版3行布局
        elif "站点人数" in str(row3) and ("🌤" in str(row2) or "尖峰" in str(row2)):
            pass  # 旧版2行布局
        else:
            errors.append(f"[全天达成] 缺少统计概要行")
        # 检查是否有表头（全员达成时不强制）
        all_achieved = False
        for r in range(3, min(ws.max_row + 1, 10)):
            row = ws.cell(r, 1).value
            if row and "全部达成" in str(row):
                all_achieved = True
                break
        if not all_achieved:
            for r in range(3, min(ws.max_row + 1, 10)):
                row = ws.cell(r, 1).value
                if row and ("排名" in str(row) or "骑手" in str(row)):
                    break
            else:
                errors.append(f"[全天达成] 未找到表头行")

    elif sn == "时段未达标":
        if "时段日报" not in str(r1) and "时段" not in str(r1):
            errors.append(f"[时段未达标] 标题不对: {r1}")

    elif sn == "档位达成":
        if "档位达成" not in str(r1):
            errors.append(f"[档位达成] 标题不对: {r1}")
        found_header = False
        for r in range(2, min(ws.max_row + 1, 10)):
            row = str(ws.cell(r, 1).value or "")
            if "档位" in row:
                found_header = True
                break
        if not found_header:
            errors.append(f"[档位达成] 未找到档位分布")

    elif sn == "重点跟进":
        if "重点跟进" not in str(r1):
            errors.append(f"[重点跟进] 标题不对: {r1}")

# ── 检查配色（表头应为深蓝 #1A365D） ──
for sn in ["全天达成", "时段未达标", "重点跟进"]:
    if sn not in actual_sheets:
        continue
    ws = wb[sn]
    c = ws.cell(1, 1)
    if c.fill and c.fill.start_color:
        rgb = str(c.fill.start_color.rgb)
        if "1A365D" not in rgb:
            errors.append(f"[{sn}] 表头配色不是 #1A365D (实际: {rgb})")

wb.close()

if errors:
    print("❌ 格式验证未通过:")
    for e in errors:
        print(f"  - {e}")
    sys.exit(1)
else:
    print("✅ 格式验证通过 — 报告符合固化标准")
    print(f"   {fp}")

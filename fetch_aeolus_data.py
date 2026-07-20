#!/usr/bin/env python3
"""
fetch_aeolus_data.py — 从 aeolus API 获取实时出勤/排班/运单数据
输出到 ~/Downloads/aeolus_api_{date}_{ts}.xlsx

用法：
    python3 fetch_aeolus_data.py
    python3 fetch_aeolus_data.py --date 2026-06-20

依赖：pip install requests openpyxl
"""
import argparse, datetime, json, os, sys, time
try:
    import requests
except ImportError:
    print("❌ 需 pip install requests"); sys.exit(1)
try:
    import openpyxl
except ImportError:
    print("❌ 需 pip install openpyxl"); sys.exit(1)

STATION = "北京-壹驰-传媒大学站_新专送"
HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Content-Type": "application/json",
    "Origin": "https://aeolus-internal.sesamepark.com",
    "Referer": "https://aeolus-internal.sesamepark.com/",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "sec-ch-ua": '"Chromium";v="136", "Not?A_Brand";v="8"',
    "sec-ch-ua-mobile": "?0", "sec-ch-ua-platform": '"macOS"',
}

def get_cookie():
    """从 ~/.aeolus_cookies/ 读取 cookie"""
    d = os.path.expanduser("~/.aeolus_cookies")
    for fn in ["aeolus_cookies.txt", "aeolus_cookies.json"]:
        fp = os.path.join(d, fn)
        if os.path.exists(fp):
            with open(fp) as f:
                txt = f.read().strip()
            if fn.endswith(".json"):
                parts = []
                for c in json.loads(txt):
                    if c.get("name") and c.get("value") and c.get("expirationDate", time.time()+1) > time.time():
                        parts.append(f"{c['name']}={c['value']}")
                return "; ".join(parts)
            return txt
    print("❌ 未找到 aeolus cookie。请先登录网页并导出 cookie 到 ~/.aeolus_cookies/")
    sys.exit(1)

def find_station(items, path=""):
    for item in items:
        label = item.get("label","")
        cp = f"{path}/{label}" if path else label
        if STATION in label:
            return item, cp
        if "children" in item:
            r = find_station(item["children"], cp)
            if r[0]: return r
    return None, None

def fetch(api, payload, timeout=30):
    url = f"https://aeolus-internal.sesamepark.com/_nuxt/api/v2/rider/{api}"
    r = requests.post(url, headers={**HEADERS, "Cookie": get_cookie()}, json=payload, timeout=timeout)
    return r.json()

def main():
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    
    # 1. 查站点 org_id
    data = fetch("tree", {"areaCode": ""})
    station, path = find_station(data.get("data", []))
    if not station:
        print("❌ 未找到站点"); sys.exit(1)
    org_id = station.get("organizationVo",{}).get("id","")
    label = station.get("label","")
    print(f"📍 {label} | org_id={org_id}")
    
    # 2. 拉取三个数据源
    print(f"\n📡 拉取 {date_str} 数据…")
    att = fetch("attendance", {"organizationId":int(org_id),"attendanceDate":date_str,"orgType":"station"}, 30)
    riders = att.get("data",{}).get("rows",[]); print(f"  考勤: {len(riders)} 人")
    sched = fetch("schedule", {"organizationId":int(org_id),"arrangedDate":date_str,"orgType":"station"}, 30)
    schedules = sched.get("data",{}).get("rows",[]); print(f"  排班: {len(schedules)} 人")
    track = fetch("tracking", {"organizationId":int(org_id),"queryDate":date_str,"type":1,"page":1,"pageSize":5000}, 60)
    tracking = track.get("data",{}).get("rows",[]); print(f"  运单: {len(tracking)} 条")
    
    # 3. 保存到 Excel
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title="实时考勤"
    if riders:
        h = list(riders[0].keys()); ws.append(h)
        for r in riders: ws.append([r.get(k,"") for k in h])
    ws2 = wb.create_sheet("排班数据")
    if schedules:
        h = list(schedules[0].keys()); ws2.append(h)
        for s in schedules: ws2.append([s.get(k,"") for k in h])
    ws3 = wb.create_sheet("运单数据")
    if tracking:
        h = list(tracking[0].keys()); ws3.append(h)
        for t in tracking: ws3.append([t.get(k,"") for k in h])
    ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    out = os.path.expanduser(f"~/Downloads/aeolus_api_{date_str}_{ts}.xlsx")
    wb.save(out); print(f"💾 {out}")
    print("✅ 完成")

if __name__ == "__main__":
    main()

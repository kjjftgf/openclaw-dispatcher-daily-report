#!/usr/bin/env python3
"""
排班实时监控-有效时长 分析脚本
按 Excel 指标表结构，计算每个骑手在每个排班时段的达标情况。
"""
import os, re, json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system('pip3 install openpyxl --break-system-packages >/dev/null 2>&1')
    import openpyxl

# 时段标准（来自 特征值 表）
SHIFT_STANDARDS = [
    {'name': '凌晨1', 'start': '00:00', 'end': '02:00', 'req_hours': 1.5,  'req_orders': 1,  'score': 1},
    {'name': '早餐1', 'start': '06:00', 'end': '08:00', 'req_hours': 1.5,  'req_orders': 1,  'score': 1},
    {'name': '早餐2', 'start': '08:00', 'end': '10:30', 'req_hours': 2.0,  'req_orders': 1,  'score': 1},
    {'name': '午高峰', 'start': '10:30', 'end': '13:30', 'req_hours': 2.75, 'req_orders': 6,  'score': 2},
    {'name': '下午茶1','start': '13:30', 'end': '15:30', 'req_hours': 1.5,  'req_orders': 2,  'score': 1},
    {'name': '下午茶2','start': '15:30', 'end': '17:30', 'req_hours': 1.5,  'req_orders': 2,  'score': 1},
    {'name': '晚高峰', 'start': '17:30', 'end': '20:00', 'req_hours': 2.25, 'req_orders': 5,  'score': 2},
    {'name': '夜宵1',  'start': '20:00', 'end': '22:00', 'req_hours': 1.5,  'req_orders': 2,  'score': 2},
    {'name': '夜宵2',  'start': '22:00', 'end': '23:59', 'req_hours': 1.5,  'req_orders': 2,  'score': 2},
]
STANDARDS_MAP = {s['name']: s for s in SHIFT_STANDARDS}

def fmt_time(h):
    """小时→友好文字"""
    if h is None or h == 0: return '0分钟'
    total_m = round(h * 60)
    hrs, mins = divmod(total_m, 60)
    if hrs == 0: return f'{mins}分钟'
    if mins == 0: return f'{hrs}小时'
    return f'{hrs}小时{mins}分钟'

def parse_time(t):
    """'7小时42分钟' / '07:42' / 秒数 → float hours"""
    if not t or t == '-' or t is None: return 0.0
    # 纯数字→秒（新格式：时长以秒存储）
    if isinstance(t, (int, float)):
        return t / 3600.0
    t = str(t).strip()
    h = 0.0; m = 0.0
    g = re.search(r'(\d+)小时', t)
    if g: h = float(g.group(1))
    g = re.search(r'(\d+)分钟', t)
    if g: m = float(g.group(1))
    if not h and not m:
        g = re.search(r'^(\d+):(\d+)$', t)
        if g: h, m = float(g.group(1)), float(g.group(2))
    return h + m / 60

def find_col(headers, targets, blacklist=None):
    """在表头中查找目标列"""
    for t in targets:
        for i, h in enumerate(headers):
            if h and t in str(h) and not (blacklist and any(b in str(h) for b in blacklist)):
                return i + 1
    return None

def parse_schedule_format(fp):
    """解析排班表格式（骑手姓名 + 排班时段字符串）"""
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Find name col and schedule col
    name_col = find_col(headers, ['骑手姓名', '姓名'])
    schedule_col = None
    for i, h in enumerate(headers):
        if h and not any(x in str(h) for x in ['id', '团队', '手机', '日期', '站点']):
            # Check if it contains period names
            pass
    # The schedule column is typically the last column with date as header
    for i, h in enumerate(headers):
        if h and ('出勤' in str(h) or '排班' in str(h) or not any(x in str(h) for x in ['id', '名称', '姓名', '手机'])):
            if i + 1 > 4:  # Not in first 4 id/name columns
                schedule_col = i + 1
    # Fallback: last non-empty column
    if not schedule_col:
        for c in range(ws.max_column, 0, -1):
            h = ws.cell(1, c).value
            if h and c > name_col:
                schedule_col = c
                break

    if not name_col or not schedule_col:
        print(f'❌ 无法识别排班表列: {headers[:15]}')
        wb.close()
        return None

    riders = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        schedule_str = ws.cell(r, schedule_col).value
        if not name or not schedule_str:
            continue
        periods = [p.strip() for p in str(schedule_str).replace('、', ',').split(',') if p.strip()]
        riders[name] = periods

    wb.close()
    return riders

def parse_attendance_format(fp):
    """解析实时出勤格式（站点/骑手/时段/在线时长/完单量）"""
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    col = {}
    for target, names in {
        '站点': ['站点名称'], '姓名': ['姓名', '骑手姓名'], '骑手id': ['骑手id', '骑手ID'],
        '时段': ['时段'], '排班': ['排班状态', '是否排班'],
        '工作': ['工作状态'],
        '配送中': ['配送中单量'],
        '全天时长': ['全天有效在线时长', '有效在线时长'], '全天完单': ['全天完单量', '完单量'],
        '时段时长': ['时段有效在线时长', '有效在线时长'], '时段完单': ['时段完单量', '完单量'],
        '在线时长': ['时段在线时长', '在线时长'],
    }.items():
        bl = ['排班'] if target == '时段' else (['目标'] if target in ('全天完单', '全天时长') else None)
        col[target] = find_col(headers, names, blacklist=bl)

    if not col['姓名'] or not col['时段']:
        wb.close()
        return None

    riders_data = []
    stations_info = {}

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col['姓名']).value
        if not name: continue

        period = ws.cell(r, col['时段']).value
        if not period or str(period).strip() == '全天': continue

        station = ws.cell(r, col['站点']).value or ''
        rider_id = ws.cell(r, col['骑手id']).value if col['骑手id'] else name
        schedule_status = str(ws.cell(r, col['排班']).value or '') if col['排班'] else ''
        work = str(ws.cell(r, col['工作']).value or '') if col['工作'] else ''

        orders_raw = ws.cell(r, col['时段完单']).value if col['时段完单'] else 0
        try: orders = float(orders_raw or 0)
        except: orders = 0

        online_raw = ws.cell(r, col['时段时长']).value if col['时段时长'] else 0
        online = parse_time(online_raw)

        online_total_raw = ws.cell(r, col['在线时长']).value if col['在线时长'] else 0
        online_total = parse_time(online_total_raw)

        delivering_raw = ws.cell(r, col['配送中']).value if col.get('配送中') else 0
        try: delivering = float(delivering_raw or 0)
        except: delivering = 0

        # 判断是否排班：直接使用是否排班列
        is_scheduled = (schedule_status.strip() == '排班')
        is_online = work.strip() == '上班' if work else (is_scheduled and online > 0)

        if not station: continue

        std = STANDARDS_MAP.get(period)
        if not std: continue

        order_ok = orders >= std['req_orders']
        time_ok = online >= std['req_hours']
        passed = 1 if (order_ok and time_ok) else 0

        result = '双不达标'
        if not is_scheduled:
            result = '时段未排班'
        elif passed:
            result = '达标'
        elif order_ok:
            result = '时长不达标'
        elif time_ok:
            result = '完单不达标'

        riders_data.append({
            'station': station, 'name': name, 'id': rider_id,
            'period': period, 'orders': orders, 'online': round(online, 2),
            'online_total': round(online_total, 2),
            'scheduled': is_scheduled, 'work_status': work, 'is_online': is_online, 'delivering': delivering,
            'req_orders': std['req_orders'], 'req_hours': std['req_hours'],
            'order_ok': order_ok, 'time_ok': time_ok,
            'passed': passed, 'result': result,
            'gap_time': round(online - std['req_hours'], 2),
            'gap_order': round(orders - std['req_orders'], 1),
        })

        if station not in stations_info:
            stations_info[station] = {'riders': set(), 'online': set(), 'periods': {}}
        if is_scheduled:
            stations_info[station]['riders'].add(name)
        if is_online:
            stations_info[station]['online'].add(name)
        if period not in stations_info[station]['periods']:
            stations_info[station]['periods'][period] = {'scheduled': 0, 'passed': 0, 'present': 0}
        if is_scheduled:
            stations_info[station]['periods'][period]['scheduled'] += 1
            if passed:
                stations_info[station]['periods'][period]['passed'] += 1
            # 'present' = has orders or online time > 0
            if orders > 0 or online > 0:
                stations_info[station]['periods'][period]['present'] += 1

    wb.close()
    return riders_data, stations_info

def print_attendance_report(filepath, riders_data, stations_info):
    """输出出勤数据报告"""
    now = datetime.now().strftime('%H:%M:%S')
    print(f"\n{'='*60}")
    print(f"📊 排班实时监控-有效时长")
    print(f"   当前时间: {now}")
    print(f"   文件: {Path(filepath).name}")
    print(f"{'='*60}")

    print(f"\n📌 【站点上线率】")
    print(f"{'站点名称':<24} {'排班骑手':>8} {'已在线':>6} {'上线率':>8}")
    print(f"{'─'*48}")
    for sname, sinfo in sorted(stations_info.items()):
        total_riders = len(sinfo['riders'])
        online_count = len(sinfo['online'])
        online_rate = online_count / total_riders * 100 if total_riders else 0
        print(f"{sname:<24} {total_riders:>8} {online_count:>6} {online_rate:>6.1f}%")

    print(f"\n📌 【各时段排班达标率】")
    print(f"{'时段':<10} {'排班':>6} {'在线':>6} {'达标':>6} {'达标率':>8}")
    print(f"{'─'*40}")
    for sname, sinfo in sorted(stations_info.items()):
        print(f"\n  {sname}")
        for pname in [s['name'] for s in SHIFT_STANDARDS]:
            p = sinfo['periods'].get(pname)
            if p and p['scheduled'] > 0:
                rate = p['passed'] / p['scheduled'] * 100
                print(f"  {pname:<8} {p['scheduled']:>6} {p['present']:>6} {p['passed']:>6} {rate:>6.1f}%")

    print(f"\n📌 【实时未达成明细】排班但未达标的骑手（仅显示午高峰、下午茶1、下午茶2）：")
    target_periods = ['午高峰','下午茶1','下午茶2']
    failures = [d for d in riders_data if d['scheduled'] and not d['passed'] and d['period'] in target_periods]
    if failures:
        print(f"{'姓名':<8} {'时段':<8} {'完单':>5} {'在线':>6} {'要求单':>4} {'要求时':>4} {'类型':<12} {'差距'}")
        print(f"{'─'*55}")
        for d in sorted(failures, key=lambda x: (x['period'], x['name'])):
            parts = []
            if not d['time_ok']: parts.append(f"差时{fmt_time(-d['gap_time'])}")
            if not d['order_ok']: parts.append(f"差单{-d['gap_order']:.0f}单")
            print(f"{d['name']:<8} {d['period']:<8} {d['orders']:>4.0f}单 {d['online']:>5.1f}h "
                  f"{d['req_orders']:>3d}单 {d['req_hours']:>3.1f}h "
                  f"{d['result']:<12} {' '.join(parts)}")
    else:
        print(f"  （当前排班时段无人未达标）")

    print(f"\n📊 汇总")
    all_scheduled = [d for d in riders_data if d['scheduled']]
    all_passed = [d for d in all_scheduled if d['passed']]
    total_all = len(all_scheduled)
    print(f"   总排班人次: {total_all} | 达标: {len(all_passed)} | 达标率: {len(all_passed)/total_all*100:.1f}%" if total_all else "")
    print(f"   未达标: {len([d for d in all_scheduled if not d['passed']])} 人次")
    print(f"   未排班: {len([d for d in riders_data if not d['scheduled']])} 人次")

def print_schedule_report(filepath, riders_schedule):
    """输出排班表报告（无出勤数据，只显示排班情况）"""
    print(f"\n{'='*60}")
    print(f"📋 排班数据 — 时排达成分板")
    print(f"   文件: {Path(filepath).name}")
    print(f"{'='*60}")

    targets = ['凌晨1','凌晨2','凌晨3','早餐1','早餐2','午高峰','下午茶1','下午茶2','晚高峰','夜宵1','夜宵2','休息']
    period_counts = {t: 0 for t in targets}
    for periods in riders_schedule.values():
        for p in periods:
            if p in period_counts:
                period_counts[p] += 1

    print(f"\n📌 【各时段排班人数】")
    for p in targets:
        if period_counts.get(p, 0) > 0:
            marker = ' ← 关注' if p in ['午高峰','下午茶1','下午茶2'] else ''
            print(f"  {p}: {period_counts.get(p, 0)}人{marker}")

    print(f"\n📌 【午高峰】排班骑手 ({period_counts.get('午高峰',0)}人)：")
    for name, periods in sorted(riders_schedule.items()):
        if '午高峰' in periods:
            extra = []
            if '下午茶1' not in periods: extra.append('缺下午茶1')
            if '下午茶2' not in periods: extra.append('缺下午茶2')
            tag = f'  ⚠️ {chr(10)+" ".join(extra)}' if extra else ''
            print(f"    {name}{tag}")

    print(f"\n📌 【下午茶1】排班骑手 ({period_counts.get('下午茶1',0)}人)：")
    for name, periods in sorted(riders_schedule.items()):
        if '下午茶1' in periods:
            extra = ' ⚠️ 缺午高峰' if '午高峰' not in periods else ''
            print(f"    {name}{extra}")

    print(f"\n📌 【下午茶2】排班骑手 ({period_counts.get('下午茶2',0)}人)：")
    for name, periods in sorted(riders_schedule.items()):
        if '下午茶2' in periods:
            extra = ' ⚠️ 缺午高峰' if '午高峰' not in periods else ''
            print(f"    {name}{extra}")

def analyze(filepath=None):
    if not filepath:
        files = sorted(Path.home().glob('Downloads/*.xlsx'),
                       key=lambda f: os.path.getmtime(f), reverse=True)
        files = [f for f in files if not f.name.startswith('.') and not f.name.startswith('~')]
        if not files:
            print("❌ 未找到 Excel 文件")
            return
        filepath = files[0]

    # 先尝试出勤格式（有在线时长+完单量）
    attendance_data = parse_attendance_format(filepath)
    if attendance_data:
        riders_data, stations_info = attendance_data
        # 基线过滤：计算每个骑手全天合计，只保留 ≥6h/18单的骑手
        BASELINE_HOURS = 6
        BASELINE_ORDERS = 18
        rider_totals = {}
        for d in riders_data:
            n = d['name']
            if n not in rider_totals:
                rider_totals[n] = {'orders': 0, 'online': 0}
            rider_totals[n]['orders'] += d['orders']
            rider_totals[n]['online'] += d['online']
        baseline_names = {n for n, t in rider_totals.items()
                          if t['orders'] >= BASELINE_ORDERS and t['online'] >= BASELINE_HOURS}
        below_names = {n for n in rider_totals if n not in baseline_names}

        # 过滤 riders_data
        riders_data_filtered = [d for d in riders_data if d['name'] in baseline_names]

        # 重建 stations_info（只保留基线骑手）
        stations_info_filtered = {}
        for d in riders_data_filtered:
            s = d['station']
            if s not in stations_info_filtered:
                stations_info_filtered[s] = {'riders': set(), 'online': set(), 'periods': {}}
            if d['scheduled']:
                stations_info_filtered[s]['riders'].add(d['name'])
            if d['is_online']:
                stations_info_filtered[s]['online'].add(d['name'])
            p = d['period']
            if p not in stations_info_filtered[s]['periods']:
                stations_info_filtered[s]['periods'][p] = {'scheduled': 0, 'passed': 0, 'present': 0}
            if d['scheduled']:
                stations_info_filtered[s]['periods'][p]['scheduled'] += 1
                if d['passed']:
                    stations_info_filtered[s]['periods'][p]['passed'] += 1
                if d['orders'] > 0 or d['online'] > 0:
                    stations_info_filtered[s]['periods'][p]['present'] += 1

        print(f"\n  🔍 基线过滤：≥{BASELINE_HOURS}h & ≥{BASELINE_ORDERS}单")
        print(f"  总骑手: {len(rider_totals)}人 | 达到基线: {len(baseline_names)}人 | 低于基线: {len(below_names)}人")
        print_attendance_report(filepath, riders_data_filtered, stations_info_filtered)
        return

    # 再尝试排班表格式（只有排班名单）
    riders_schedule = parse_schedule_format(filepath)
    if riders_schedule:
        print_schedule_report(filepath, riders_schedule)
        return

    print(f"❌ 无法识别文件格式")

if __name__ == '__main__':
    import sys
    analyze(sys.argv[1] if len(sys.argv) > 1 else None)
